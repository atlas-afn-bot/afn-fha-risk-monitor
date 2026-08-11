#!/usr/bin/env python3
"""
Build a monthly FHA Risk Monitor snapshot JSON from the 6 Excel source files.

Usage:
    python3 scripts/build-snapshot.py <period>

where <period> is a YYYY-MM key matching the folder under data/source/.

Reads:
    data/source/<period>/HUD Total Compare Ratios *.xlsx
    data/source/<period>/HOC Compare Ratios - *.xlsx
    data/source/<period>/HUD Field Offices - *.xlsx
    data/source/<period>/HUD Branches - *.xlsx
    data/source/<period>/NW Data *.xlsx
    data/source/<period>/Neighborhood Watch Report <period> *Enc Data.xlsx

Writes:
    public/data/snapshots/<period>.json
    public/data/snapshots/index.json       (appended / updated)

The script is idempotent — rerunning replaces the output cleanly.
"""
from __future__ import annotations

import argparse
import datetime as dt
import glob
import json
import math
import os
import sys
from collections import OrderedDict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl
import pandas as pd

SCHEMA_VERSION = 1
SCRIPT_VERSION = "1.0"
REPO_ROOT = Path(__file__).resolve().parent.parent
# Environment overrides so the automated snapshot pipeline (container)
# can point the script at a working tree it built at runtime, without
# having to duplicate the whole repo under /opt/repo. The default
# behavior (local dev, run from a repo checkout) is unchanged.
_env_source = os.environ.get("SNAPSHOT_SOURCE_ROOT")
_env_out = os.environ.get("SNAPSHOT_OUTPUT_DIR")
SOURCE_ROOT = Path(_env_source) if _env_source else REPO_ROOT / "data" / "source"
SNAPSHOT_DIR = Path(_env_out) if _env_out else REPO_ROOT / "public" / "data" / "snapshots"

# Canonical HUD Office → HOC map (from db/migrations/001_initial_schema.sql)
HUD_OFFICE_HOC: Dict[str, str] = {
    # Atlanta
    "Atlanta": "Atlanta", "Birmingham": "Atlanta", "Caribbean": "Atlanta",
    "Columbia": "Atlanta", "Coral Gables": "Atlanta", "Greensboro": "Atlanta",
    "Jackson": "Atlanta", "Jacksonville": "Atlanta", "Knoxville": "Atlanta",
    "Louisville": "Atlanta", "Memphis": "Atlanta", "Miami": "Atlanta",
    "Nashville": "Atlanta", "Orlando": "Atlanta", "San Juan": "Atlanta",
    "Tampa": "Atlanta",
    # Denver
    "Albuquerque": "Denver", "Casper": "Denver", "Dallas": "Denver",
    "Denver": "Denver", "Des Moines": "Denver", "Fargo": "Denver",
    "Fort Worth": "Denver", "Helena": "Denver", "Houston": "Denver",
    "Kansas City": "Denver", "Little Rock": "Denver", "Lubbock": "Denver",
    "Minneapolis": "Denver", "New Orleans": "Denver", "Oklahoma City": "Denver",
    "Omaha": "Denver", "Rapid City": "Denver", "Salt Lake City": "Denver",
    "San Antonio": "Denver", "Shreveport": "Denver", "Sioux Falls": "Denver",
    "Springfield": "Denver", "St. Louis": "Denver", "Tulsa": "Denver",
    "Wichita": "Denver",
    # Philadelphia
    "Albany": "Philadelphia", "Baltimore": "Philadelphia", "Bangor": "Philadelphia",
    "Boston": "Philadelphia", "Buffalo": "Philadelphia", "Burlington": "Philadelphia",
    "Charleston": "Philadelphia", "Charlotte": "Philadelphia", "Chicago": "Philadelphia",
    "Cincinnati": "Philadelphia", "Cleveland": "Philadelphia", "Columbus": "Philadelphia",
    "Detroit": "Philadelphia", "Flint": "Philadelphia", "Grand Rapids": "Philadelphia",
    "Hartford": "Philadelphia", "Indianapolis": "Philadelphia", "Manchester": "Philadelphia",
    "Milwaukee": "Philadelphia", "Newark": "Philadelphia", "New York": "Philadelphia",
    "Philadelphia": "Philadelphia", "Pittsburgh": "Philadelphia", "Providence": "Philadelphia",
    "Richmond": "Philadelphia", "Washington, DC": "Philadelphia",
    # Santa Ana
    "Anchorage": "Santa Ana", "Boise": "Santa Ana", "Fresno": "Santa Ana",
    "Honolulu": "Santa Ana", "Las Vegas": "Santa Ana", "Los Angeles": "Santa Ana",
    "Phoenix": "Santa Ana", "Portland": "Santa Ana", "Reno": "Santa Ana",
    "Sacramento": "Santa Ana", "San Diego": "Santa Ana", "San Francisco": "Santa Ana",
    "Santa Ana": "Santa Ana", "Seattle": "Santa Ana", "Spokane": "Santa Ana",
    "Tucson": "Santa Ana",
}

MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _clean_num(v: Any) -> Optional[float]:
    """Return a finite float or None — normalizes Excel empties / strings."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None  # avoid True/False sneaking in as 1/0
    if isinstance(v, (int, float)):
        if isinstance(v, float) and math.isnan(v):
            return None
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _clean_int(v: Any) -> Optional[int]:
    f = _clean_num(v)
    if f is None:
        return None
    return int(f)


def _clean_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    # Reject pandas NaN floats (they stringify to "nan")
    if isinstance(v, float) and math.isnan(v):
        return None
    s = str(v).strip()
    if not s:
        return None
    # Reject literal "nan"/"NaT" strings produced by pandas coercion
    if s.lower() in {"nan", "nat", "none"}:
        return None
    return s


def _case_norm(v: Any) -> Optional[str]:
    """Normalize FHA case number so Encompass (010-1234567) matches HUD (010-1234567)."""
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    # Strip trailing spaces / zero-pad nothing — just return as-is uppercased
    return s.upper()


# BIFF (real Excel 97-2003 .xls) magic number — D0 CF 11 E0 A1 B1 1A E1.
_BIFF_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _xls_to_xlsx(xls_path: Path) -> Path:
    """Convert a legacy ``.xls`` file to ``.xlsx`` in the same directory.

    Downstream readers use ``openpyxl``/``pd.read_excel(engine="openpyxl")``,
    which cannot open ``.xls``. RPA occasionally uploads ``.xls`` — and in
    HUD's case, "``.xls``" is often a **tab-delimited text file** with an
    ``.xls`` extension (HUD Neighborhood Watch's "Export to Excel" quirk),
    not a real BIFF workbook.

    This helper detects which flavor we have and normalizes both to a proper
    ``.xlsx`` so the rest of the pipeline stays engine-agnostic.

    The converted file is cached next to the source and reused on re-runs.
    Requires ``xlrd`` (pinned in ``scripts/requirements.txt``) for real BIFF.
    """
    from openpyxl import Workbook

    out_path = xls_path.with_suffix(".xlsx")
    if out_path.exists() and out_path.stat().st_mtime >= xls_path.stat().st_mtime:
        return out_path

    with open(xls_path, "rb") as fh:
        head = fh.read(8)

    if head.startswith(_BIFF_MAGIC):
        # Real Excel 97-2003 workbook — use xlrd.
        import xlrd  # lazy import
        book = xlrd.open_workbook(str(xls_path), formatting_info=False)
        wb = Workbook()
        default = wb.active
        wb.remove(default)
        for sheet_name in book.sheet_names():
            sh = book.sheet_by_name(sheet_name)
            ws = wb.create_sheet(title=sheet_name[:31] or "Sheet1")
            for r in range(sh.nrows):
                row_vals = []
                for c in range(sh.ncols):
                    cell = sh.cell(r, c)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            row_vals.append(
                                xlrd.xldate.xldate_as_datetime(cell.value, book.datemode)
                            )
                        except Exception:
                            row_vals.append(cell.value)
                    elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                        row_vals.append(bool(cell.value))
                    elif cell.ctype == xlrd.XL_CELL_ERROR:
                        row_vals.append(None)
                    else:
                        row_vals.append(cell.value)
                ws.append(row_vals)
        wb.save(str(out_path))
        return out_path

    # Not BIFF — assume a delimited text file (HUD NW "Export to Excel" ships
    # tab-separated text with an .xls extension). Read it as text and
    # re-shape row-by-row into an xlsx. Rows use CR / CRLF / LF terminators.
    raw_bytes = xls_path.read_bytes()
    for enc in ("utf-8", "utf-16", "cp1252", "latin-1"):
        try:
            text = raw_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:  # pragma: no cover — latin-1 above should always succeed
        text = raw_bytes.decode("latin-1", errors="replace")

    # Normalize line endings (HUD uses CR-only for some files, CRLF for others).
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = text.split("\n")

    # Sniff delimiter: HUD ships \t, but some exports use commas. Pick whichever
    # produces more columns on the first non-empty line.
    sample = next((l for l in lines if l.strip()), "")
    delim = "\t" if sample.count("\t") >= sample.count(",") else ","

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for line in lines:
        if not line and not lines:  # skip a single trailing empty split
            continue
        ws.append(line.split(delim))
    wb.save(str(out_path))
    return out_path


# ---------------------------------------------------------------------------
# Slot alias table — canonical + RPA + legacy naming conventions.
#
# Callers pass a slot key; ``_find_source`` walks the candidate globs in order
# (first hit wins, with an .xlsx-over-.xls tiebreaker). Adding a new naming
# convention means appending one string here — nothing else changes.
# ---------------------------------------------------------------------------
SLOT_ALIAS_TABLE: Dict[str, List[str]] = {
    "hud_total_compare_ratios": [
        "HUD Total Compare Ratios*.xlsx",     # legacy / manual naming
        "HUD_National_Totals*.xlsx",          # 2026-06+ RPA naming
        "HUD_Total_Compare_Ratio*.xlsx",
        "HUD Total Compare Ratios*.xls",
        "HUD_National_Totals*.xls",
        "HUD_Total_Compare_Ratio*.xls",       # 2026-05 RPA naming (.xls)
    ],
    "hoc_compare_ratios": [
        "HOC Compare Ratios*.xlsx",
        "HOCs*.xlsx",                         # 2026-06+ RPA naming
        "HOC_Compare_Ratios*.xlsx",
        "HOC Compare Ratios*.xls",
        "HOCs*.xls",
        "HOC_Compare_Ratios*.xls",            # 2026-05 RPA naming (.xls)
    ],
    "hud_field_offices": [
        "HUD Field Offices*.xlsx",
        "HUD_Office*.xlsx",                   # 2026-06+ RPA naming (singular)
        "HUD_Offices*.xlsx",
        "HUD Field Offices*.xls",
        "HUD_Office*.xls",
        "HUD_Offices*.xls",                   # 2026-05 RPA naming (.xls)
    ],
    "hud_branches": [
        "HUD Branches*.xlsx",
        "Branches*.xlsx",                     # 2026-06+ RPA naming
        "HUD_Branches*.xlsx",
        "HUD Branches*.xls",
        "Branches*.xls",
        "HUD_Branches*.xls",                  # 2026-05 RPA naming (.xls)
    ],
    "nw_data": [
        "NW Data*.xlsx",
        "NW_Data*.xlsx",                      # 2026-06+ RPA naming
        "NW Data*.xls",
        "NW_Data*.xls",                       # 2026-05 RPA naming (.xls)
    ],
    "hud_total_population": [
        "NW Total Population*.xlsx",
        "NW_Total_Population*.xlsx",
        "NW Total Population*.xls",
        "NW_Total_Population*.xls",
    ],
}


def _find_source(period: str, slot_or_pattern) -> Path:
    """Locate a source Excel under ``data/source/{period}/``.

    ``slot_or_pattern`` may be one of:

    * a **slot key** from :data:`SLOT_ALIAS_TABLE` (preferred),
    * a **list of glob patterns** (first hit wins), or
    * a **single glob pattern** (legacy shape — kept for callers that pass a raw glob).

    Selection rules when multiple candidates match:

    1. Prefer ``.xlsx`` over ``.xls`` (openpyxl-native beats legacy BIFF).
    2. Then most-recent ``mtime``.

    ``.xls`` matches are transparently converted to ``.xlsx`` in-place
    (see :func:`_xls_to_xlsx`) so downstream readers never see BIFF files.

    If nothing matches, raises ``FileNotFoundError`` listing **every** glob
    that was tried — not just the first — so operators can see what the
    script expected vs. what the RPA actually uploaded.
    """
    base = SOURCE_ROOT / period

    if isinstance(slot_or_pattern, str) and slot_or_pattern in SLOT_ALIAS_TABLE:
        patterns = list(SLOT_ALIAS_TABLE[slot_or_pattern])
        label = slot_or_pattern
    elif isinstance(slot_or_pattern, (list, tuple)):
        patterns = list(slot_or_pattern)
        label = None
    else:
        patterns = [str(slot_or_pattern)]
        label = None

    # (pattern, path) tuples for every match across every glob.
    hits: List[Tuple[str, Path]] = []
    for pat in patterns:
        for p in sorted(base.glob(pat)):
            hits.append((pat, p))

    if not hits:
        listing = "\n    ".join(patterns)
        prefix = f"{label} slot" if label else "any of the expected patterns"
        raise FileNotFoundError(
            f"No file matches {prefix} in {base}. Tried:\n    {listing}"
        )

    # Prefer .xlsx over .xls; then most-recent mtime.
    def _rank(entry: Tuple[str, Path]) -> Tuple[int, float]:
        _pat, path = entry
        ext_rank = 0 if path.suffix.lower() == ".xlsx" else 1  # 0 wins
        return (ext_rank, -path.stat().st_mtime)

    hits.sort(key=_rank)
    matched_pattern, chosen = hits[0]
    print(f"  _find_source: matched {chosen.name!r} via glob {matched_pattern!r}")

    if chosen.suffix.lower() == ".xls":
        converted = _xls_to_xlsx(chosen)
        print(f"  _find_source: converted .xls → .xlsx  {chosen.name!r} → {converted.name!r}")
        return converted
    return chosen


def _title_case_office(name: str) -> str:
    """Normalize HUD office name for HOC lookup (HUD exports are padded uppercase)."""
    s = name.strip().title()
    # HUD exports "WASHINGTON, DC" etc. — preserve the DC
    s = s.replace("Dc", "DC").replace("Usa", "USA")
    return s


def _match_hoc(office_name: str) -> Optional[str]:
    return HUD_OFFICE_HOC.get(_title_case_office(office_name))


def _parse_performance_period(raw_cell: str) -> Tuple[str, str]:
    """Extract an ISO date and window label from the Performance Period cell.

    Accepts strings like:
        "Performance Period - 02/28/2026"
        "Data shown includes all insured single family loans with beginning amortization date between March 1, 2024 and February 28, 2026"
    """
    raw = raw_cell.strip()
    # MM/DD/YYYY
    import re
    m = re.search(r"(\d{2})/(\d{2})/(\d{4})", raw)
    if m:
        mm, dd, yyyy = m.groups()
        iso = f"{yyyy}-{mm}-{dd}"
        return iso, raw
    # "between X and Y"
    m = re.search(r"between\s+(.+?)\s+and\s+(.+?)$", raw, re.IGNORECASE)
    if m:
        end = m.group(2).strip().rstrip(".")
        # Expect "February 28, 2026"
        m2 = re.match(r"(\w+)\s+(\d+),?\s+(\d{4})", end)
        if m2:
            mon, day, yr = m2.group(1), m2.group(2), m2.group(3)
            mnum = MONTH_NAMES.index(mon) + 1 if mon in MONTH_NAMES else 1
            return f"{yr}-{mnum:02d}-{int(day):02d}", f"{m.group(1).strip()} — {end}"
    return "", raw


# ─────────────────────────────────────────────────────────────────────────────
# Readers — each returns a list-of-dicts in the snapshot shape
# ─────────────────────────────────────────────────────────────────────────────

def read_compare_ratios_total(path: Path) -> Tuple[List[dict], str, str]:
    """Return (rows, iso_date, window_label) from the Total Compare Ratios file."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    perf_date, perf_label = "", ""
    for r in rows:
        if not r or not r[0]:
            continue
        text = str(r[0])
        low = text.lower()
        if "performance period" in low or "amortization date between" in low:
            iso, lbl = _parse_performance_period(text)
            if iso and not perf_date:
                perf_date = iso
            # The "between March 1, 2024 and February 28, 2026" row yields the
            # richest window label; prefer it when available.
            if lbl and ("—" in lbl or " and " in lbl.lower()) and not perf_label:
                perf_label = lbl

    # Header is row 8 (index 7), data at row 9 (index 8)
    if len(rows) < 9:
        raise RuntimeError(f"{path}: expected 9+ rows, got {len(rows)}")
    data = rows[8]

    total = {
        "scope": "total",
        "compare_ratio": _clean_num(data[0]),
        "mix_adjusted_sdq": _clean_num(data[26]) if len(data) > 26 else None,
        "fha_benchmark_sdq": _clean_num(data[28]) if len(data) > 28 else None,
        "supplemental_metric": _clean_num(data[25]) if len(data) > 25 else None,
        "loans_count": _clean_int(data[3]),
        "delinquent_count": _clean_int(data[4]),
    }
    retail = {
        "scope": "retail",
        "compare_ratio": _clean_num(data[1]),
        "mix_adjusted_sdq": None,
        "fha_benchmark_sdq": None,
        "supplemental_metric": None,
        "loans_count": _clean_int(data[7]),
        "delinquent_count": _clean_int(data[9]),
    }
    sponsor = {
        "scope": "sponsor",
        "compare_ratio": _clean_num(data[2]),
        "mix_adjusted_sdq": None,
        "fha_benchmark_sdq": None,
        "supplemental_metric": None,
        "loans_count": _clean_int(data[12]),
        "delinquent_count": _clean_int(data[14]),
    }
    return [total, retail, sponsor], perf_date, perf_label


def read_compare_ratios_hoc(path: Path) -> List[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    out: List[dict] = []
    # Header row 8 (index 7); data rows start at index 8
    for r in rows[8:]:
        if not r or not r[0]:
            continue
        name = str(r[0]).strip().title()
        if name not in {"Atlanta", "Denver", "Philadelphia", "Santa Ana"}:
            continue
        out.append({
            "hoc_name": name,
            "compare_ratio": _clean_num(r[1]),
            "retail_ratio": _clean_num(r[2]),
            "sponsor_ratio": _clean_num(r[3]),
            "mix_adjusted_sdq": None,     # not in HOC file
            "fha_benchmark_sdq": None,
            "supplemental_metric": None,
            "loans_count": _clean_int(r[4]),
            "delinquent_count": _clean_int(r[5]),
        })
    return out


def read_compare_ratios_hud_office(path: Path) -> List[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find header
    header_idx = -1
    for i, r in enumerate(rows[:20]):
        if r and r[0] and "HUD OFFICE" in str(r[0]).upper() and len(r) > 1 \
                and "COMPARE RATIO" in str(r[1] or "").upper():
            header_idx = i
            break
    if header_idx < 0:
        raise RuntimeError(f"{path}: no header row found")

    out: List[dict] = []
    for r in rows[header_idx + 1:]:
        if not r or not r[0]:
            continue
        name = str(r[0]).strip()
        upper = name.upper()
        if upper.startswith(("REPORT", "OUTPUT", "LOAN TYPE", "DATA SHOWN")):
            continue
        canonical = _title_case_office(name)
        out.append({
            "hud_office": canonical,
            "hoc": _match_hoc(canonical),
            "retail_branches_count": _clean_int(r[7]),
            "sponsored_branches_count": _clean_int(r[12]),
            "compare_ratio": _clean_num(r[1]),
            "retail_ratio": _clean_num(r[2]),
            "sponsor_ratio": _clean_num(r[3]),
            "loans_count": _clean_int(r[4]),
            "delinquent_count": _clean_int(r[5]),
            "retail_loans": _clean_int(r[8]),
            "retail_delinquent": _clean_int(r[10]),
            "sponsored_loans": _clean_int(r[13]),
            "sponsored_delinquent": _clean_int(r[15]),
            "hud_office_dq_pct": _clean_num(r[19]),
            "area_retail_dq_pct": _clean_num(r[22]),
            "area_sponsored_dq_pct": _clean_num(r[25]),
            "mix_adjusted_sdq": _clean_num(r[27]) if len(r) > 27 else None,
            "fha_benchmark_sdq": _clean_num(r[29]) if len(r) > 29 else None,
            "supplemental_metric": _clean_num(r[26]) if len(r) > 26 else None,
        })
    return out


def read_hud_total_population(path: Path) -> Dict[str, dict]:
    """Read HUD's Neighborhood Watch total-population xlsx (all loans, not just SDQ).

    HUD delivers this on request as `NW Total Population <M.D.YY>.xlsx`. It has
    the same schema as `NW Data *.xlsx` (row 8 header, data rows 9+). Each row
    is one FHA-insured loan HUD attributes to AFN.

    Returns a dict keyed by FHA Case Number, with:
      hud_channel        : 'Retail' if Originating ID populated, else 'Wholesale'
      hud_orig_id        : HUD Originating ID (10-digit FHA branch ID) or ''
      hud_sponsor_id     : HUD Sponsor ID or ''
      hud_fha_ins_stat   : 'A' (active) / 'C' (claim) / etc.

    Used to filter the Encompass loan set to HUD's authoritative population
    so the snapshot's total loan count matches HUD Compare Ratios exactly.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # HUD NW files: header row is row 8 (0-indexed), data starts row 9.
    out: Dict[str, dict] = {}
    for r in rows[9:]:
        if not r or r[7] is None:
            continue
        case = str(r[7]).strip()
        if not case:
            continue
        orig = str(r[0] or "").strip()
        spr = str(r[1] or "").strip()
        out[case] = {
            "hud_channel": "Retail" if orig else ("Wholesale" if spr else None),
            "hud_orig_id": orig,
            "hud_sponsor_id": spr,
            "hud_fha_ins_stat": str(r[8] or "").strip() if len(r) > 8 else "",
        }
    return out


def read_compare_ratios_branch(path: Path) -> List[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Header row 8 (index 7) starts with "Retail Branch"
    header_idx = -1
    for i, r in enumerate(rows[:20]):
        if r and r[0] and "RETAIL BRANCH" in str(r[0]).upper():
            header_idx = i
            break
    if header_idx < 0:
        raise RuntimeError(f"{path}: no branch header found")

    out: List[dict] = []
    for r in rows[header_idx + 1:]:
        if not r or r[0] is None:
            continue
        nmls = _clean_str(r[0])
        if not nmls:
            continue
        # Branch NMLS IDs are 10-digit numeric strings (e.g. '1835202534').
        # Reject any row where col 0 isn't a clean numeric id — that filters
        # out the sheet footer ("(K) = Cumulative...") and any stray rows.
        nmls_digits = nmls.replace(" ", "")
        if not nmls_digits.isdigit():
            continue
        # Approval status must be 'A' or 'T'; anything else means it's not a
        # real branch row.
        approval = _clean_str(r[1])
        if approval not in ("A", "T"):
            continue
        out.append({
            "nmls_id": nmls_digits,
            "branch_name": None,  # not present in the drill-down sheet
            "hud_office": None,   # needs separate NW Branch by Office sheet (phase 2)
            "approval_status": approval,
            "loans_underwritten": _clean_int(r[3]),
            "delinquency_rate": _clean_num(r[8]),
            "compare_ratio": _clean_num(r[2]),
        })
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Encompass + NW Data 2 (loans)
# ─────────────────────────────────────────────────────────────────────────────

def _read_nw_data2(path: Path) -> pd.DataFrame:
    """Read `NW Data 2.28.26.xlsx` as a DataFrame keyed by Case Number."""
    # Header is row 9 (index 8); data starts row 10
    df = pd.read_excel(path, sheet_name=0, header=8, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    if "Case Number" not in df.columns:
        raise RuntimeError(f"{path}: expected 'Case Number' column, got {list(df.columns)[:5]}")
    df["Case Number"] = df["Case Number"].astype(str).str.strip().str.upper()
    return df


def _read_encompass(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=0, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    df["Case #"] = df["Case #"].astype(str).str.strip().str.upper()
    return df


def _to_bool(val: Any) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        if isinstance(val, float) and math.isnan(val):
            return False
        return val != 0
    s = str(val).strip().lower()
    return s in {"yes", "y", "true", "1"}


def _normalize_program(raw: str) -> str:
    s = (raw or "").strip().lower()
    if not s:
        return "Non-DPA"
    if "boost" in s:
        return "Boost"
    if "arrive" in s or "aurora" in s:
        return "Arrive/Aurora"
    return raw.strip()


def _pick(row: dict, *keys: str, default=None):
    for k in keys:
        if k in row:
            v = row[k]
            if v is None:
                continue
            if isinstance(v, float) and math.isnan(v):
                continue
            if isinstance(v, str) and not v.strip():
                continue
            return v
    return default


def _fails_enhanced_guidelines(fico: float, units: int, aus: str, reserves: float,
                               gift_amount: float, pay_shock_over_100: bool,
                               is_boost: bool) -> bool:
    if not is_boost:
        return False
    aus_upper = (aus or "").upper().strip()
    is_manual = "MANUAL" in aus_upper
    is_auto = aus_upper in ("DU", "LP") or (aus_upper != "" and not is_manual)

    if fico and fico < 640:
        return True
    if units and units >= 3:
        return True

    has_gift = (gift_amount or 0) > 0
    if is_auto:
        if fico < 680:
            if (reserves or 0) < 2:
                return True
            if has_gift:
                return True
        else:
            if (reserves or 0) < 1:
                return True
    if is_manual:
        if fico >= 680:
            if (reserves or 0) < 1:
                return True
            if has_gift:
                return True
        else:
            if (reserves or 0) < 3:
                return True
            if has_gift:
                return True
            if pay_shock_over_100:
                return True
    return False


def build_loans(enc_path: Path, nw2_path: Path,
                hud_office_lookup: Dict[str, dict]) -> List[dict]:
    print(f"  Loading Encompass: {enc_path.name}")
    enc = _read_encompass(enc_path)
    print(f"    {len(enc):,} rows × {len(enc.columns)} cols")

    print(f"  Loading NW Data 2: {nw2_path.name}")
    nw2 = _read_nw_data2(nw2_path)
    print(f"    {len(nw2):,} rows × {len(nw2.columns)} cols")

    # Left-join on Case #
    nw2_small = nw2[[c for c in [
        "Case Number", "FHA Ins Stat", "Term", "Liv Units", "Loan Purpose",
        "Mortgage Amount", "Interest Rate", "Front Ratio", "Back Ratio",
        "Loan To Value Ratio", "Credit Score", "Underwriter Name", "Unwtr ID",
        "Seriously Delinquent", "Oldest Unpaid Installment Due Date",
        "Number of Months Delinquent", "Delinquent Status", "Delinquent Reason",
        "Loan Officer NMLS ID", "Indem",
        # NW Data extension fields (sponsor / TPO, gift letter, census,
        # underwriter review, indemnification). Merged onto every loan
        # so downstream rollups can be computed from the canonical loan
        # array.
        "Sponsor ID", "Sponsored Originator Name",
        "Sponsored Originator EIN ID (last 4 digits)",
        "Sponsored Originator NMLS ID",
        "Gift Ltr Amt", "Gift Ltr Source",
        "Census Tract", "Underserved Indicator",
        "Unwtr Rvw Appr", "Unwtr Mort Cr Rtng",
        "Delinquent Status Date",
        "Payments before First 90 Day Delinquent Reported",
    ] if c in nw2.columns]].copy()
    nw2_small = nw2_small.rename(columns={"Case Number": "Case #"})
    merged = enc.merge(nw2_small, how="left", on="Case #", suffixes=("", "_nw"))
    print(f"    joined: {len(merged):,} rows")

    loans: List[dict] = []
    for idx, row in merged.iterrows():
        row = row.to_dict()

        fico = _clean_num(row.get("FICO")) or 0
        units_val = _clean_int(row.get("Subject Property # Units")) or 0
        front_dti = _clean_num(row.get("Top Ratio"))
        back_dti = _clean_num(row.get("Bottom Ratio"))
        ltv = _clean_num(row.get("LTV"))
        reserves_months = _clean_num(row.get("Reserves")) or 0
        gift_amount = _clean_num(row.get("Gift Fund Amount")) or 0
        payment_shock = _clean_num(row.get("Payment Shock"))
        pay_shock_over_100_flag = str(row.get("Pay Shock > 100") or "").strip().lower() == "yes"

        raw_program = str(row.get("Loan Program") or "")
        is_dpa = "DPA" in raw_program.upper()
        dpa_program_norm = _normalize_program(str(row.get("DPA Program") or ""))
        is_boost = is_dpa and "boost" in str(row.get("DPA Program") or "").lower()

        dq_yes = str(row.get("DQ") or "").strip().lower() == "yes"
        sdq_yes = str(row.get("Seriously Delinquent") or "").strip().lower() == "yes"
        status_code = str(row.get("HUD Status Code") or "").strip().upper()
        is_claim = dq_yes and status_code in {"C", "CLAIM", "9"}

        channel_raw = str(row.get("Loan Info Channel") or "").lower()
        if "retail" in channel_raw:
            channel = "Retail"
        elif "wholesale" in channel_raw:
            channel = "Wholesale"
        else:
            channel = None

        hud_office = _clean_str(row.get("HUD Office"))
        hud_office_norm = _title_case_office(hud_office) if hud_office else None
        hoc = _match_hoc(hud_office_norm) if hud_office_norm else _clean_str(row.get("HOC"))

        variable_pct = _clean_num(row.get("Variable Income %")) or 0
        has_variable = str(row.get("Variable Income (Y/N)") or "").strip().lower() == "y"
        has_super_var = variable_pct > 25

        manufactured_flag = _to_bool(row.get("Manufactured")) or _to_bool(row.get("RISK: Manufactured Home"))
        manual_uw = "MANUAL" in (str(row.get("Underwriting Risk Assess Type") or "")).upper()

        hud_def = _to_bool(row.get("HUDVA Condition 1")) or _to_bool(row.get("HUDVA UW Condition 1"))
        gift_grant = (_clean_num(row.get("Total Gifts and Grants")) or 0) > 0 \
            or str(row.get("Gift or Grant (Y/N)") or "").strip().lower() == "y"

        non_owner = _to_bool(row.get("Non-Owner Occupied Borrower"))

        # Risk indicator bit flags
        has_sub_620 = fico > 0 and fico < 620
        has_super_29_dti = (front_dti or 0) > 29
        has_super_50_dti = (back_dti or 0) > 50
        has_super_90_ltv = (ltv or 0) > 90
        has_super_95_ltv = (ltv or 0) > 95

        # Use RPA-supplied Risk Indicator Count when available; else recompute
        risk_count = _clean_int(row.get("Risk Indicator Count"))
        if risk_count is None:
            risk_count = sum(int(b) for b in (
                has_sub_620, has_super_29_dti, has_super_50_dti,
                has_super_90_ltv, has_super_95_ltv, is_dpa,
                manufactured_flag, has_variable, has_super_var,
                non_owner, manual_uw, hud_def, gift_grant,
            ))

        fails_eg = _fails_enhanced_guidelines(
            fico=fico, units=units_val, aus=str(row.get("Underwriting Risk Assess Type") or ""),
            reserves=reserves_months, gift_amount=gift_amount,
            pay_shock_over_100=pay_shock_over_100_flag, is_boost=is_boost,
        )

        # Occupancy / property
        occupancy = _clean_str(row.get("Occupancy Borr Pair 1"))
        property_type = _clean_str(row.get("Property Type Master")) or _clean_str(row.get("HUD 92900 LT Subject Property Type"))

        loan_id = str(row.get("Loan Number") or row.get("Case #") or f"row-{idx}").strip()

        loans.append({
            "loan_id": loan_id,
            "fha_case_number": _clean_str(row.get("Case #")),

            "loan_officer": _clean_str(row.get("Loan Officer - Retail")),
            # Prefer Encompass's LO Employee ID — AFN's canonical LO id,
            # populated for every loan. `Loan Officer NMLS ID` comes from
            # the NW Data 2 join (only ~5% of loans = delinquent ones)
            # and would shatter each LO into many single-loan buckets.
            "lo_nmls_id": _clean_str(row.get("LO Employee ID")) or _clean_str(row.get("Loan Officer NMLS ID")),
            "branch_nmls_id": _clean_str(row.get("Org ID")) or _clean_str(row.get("Broker Lender Company ID")),
            "tpo_broker": _clean_str(row.get("TPO Broker")),
            "broker": _clean_str(row.get("Broker")),
            "branch_name": _clean_str(row.get("Branch Name")),
            "branch_name_retail": _clean_str(row.get("Branch Name - Retail")),
            "hud_office": hud_office_norm,
            "hoc": hoc,
            "channel": channel,

            "dpa_program": dpa_program_norm,
            "dpa_name": _clean_str(row.get("DPA Name")),
            "dpa_investor": _clean_str(row.get("DPA Investor")),
            "investor_name": _clean_str(row.get("Investor Name")),
            "loan_purpose": _clean_str(row.get("Loan Purpose")),

            "fico_score": int(fico) if fico else None,
            "front_dti": front_dti,
            "back_dti": back_dti,
            "ltv": ltv,
            "loan_amount": _clean_num(row.get("Total Loan Amount")) or _clean_num(row.get("Mortgage Amount")),
            "source_of_funds": _clean_str(row.get("Source of Funds")),
            "employment_type": _clean_str(row.get("Self Employed (Y/N)")),
            "aus": _clean_str(row.get("Underwriting Risk Assess Type")),
            "units": units_val or None,
            "property_type": property_type,
            "occupancy": occupancy,

            "delinquent_status_code": _clean_str(row.get("HUD Status Code")),
            "delinquent_status": _clean_str(row.get("Status")) or _clean_str(row.get("Delinquent Status")),
            "months_delinquent": _clean_int(row.get("Number of Months Delinquent")),
            "oldest_unpaid_installment": _iso_date(row.get("Oldest Unpaid Installment Due Date")),
            "fha_ins_stat": _clean_str(row.get("FHA Ins Stat")),

            "has_sub_620": bool(has_sub_620),
            "has_super_29_dti": bool(has_super_29_dti),
            "has_super_50_dti": bool(has_super_50_dti),
            "has_super_90_ltv": bool(has_super_90_ltv),
            "has_super_95_ltv": bool(has_super_95_ltv),
            "has_dpa": bool(is_dpa),
            "has_manufactured": bool(manufactured_flag),
            "has_variable_income": bool(has_variable),
            "has_super_variable_income": bool(has_super_var),
            "has_non_owner_occupied": bool(non_owner),
            "has_manual_uw": bool(manual_uw),
            "has_hud_deficiency": bool(hud_def),
            "has_gift_grant": bool(gift_grant),
            "risk_indicator_count": int(risk_count) if risk_count is not None else 0,

            "is_delinquent": bool(dq_yes),
            "is_seriously_delinquent": bool(sdq_yes or dq_yes),
            "is_claim": bool(is_claim),

            "loan_program_raw": _clean_str(raw_program),
            "ltv_group": _clean_str(row.get("LTV Group")),
            "fthb": _clean_str(row.get("FTHB")),
            "dti_back_end_group": _clean_str(row.get("DTI Back End Group")),
            "payment_shock_group": _clean_str(row.get("Payment Shock Group")),
            "source_of_funds_group": _clean_str(row.get("Source of Funds Group")),
            "reserves_group": _clean_str(row.get("Reserves Group")),
            "gift_grant_group": _clean_str(row.get("% Funds from Gift or Grant Group")),
            "reserves_months": reserves_months if reserves_months else None,
            "gift_fund_amount": gift_amount if gift_amount else None,
            "payment_shock": payment_shock,
            "pay_shock_over_100": _clean_str(row.get("Pay Shock > 100")),
            "is_boost": bool(is_boost),
            "fails_enhanced_guidelines": bool(fails_eg),
            "hud_office_compare_ratio": _clean_num(row.get("HUD Office Compare Ratio")),
            "program_type": "DPA" if is_dpa else "Standard",

            # ── NW Data extension fields (additive; may be None when the
            # Encompass row didn't match a NW Data 2 row on Case #) ──
            "underwriter_name": _clean_str(row.get("Underwriter Name")),
            "underwriter_id": _clean_str(row.get("Unwtr ID")),
            "underwriter_review_approval": _clean_str(row.get("Unwtr Rvw Appr")),
            "underwriter_mortgage_credit_rating": _clean_str(row.get("Unwtr Mort Cr Rtng")),
            "sponsor_id": _clean_str(row.get("Sponsor ID")),
            "sponsor_originator_name": _clean_str(row.get("Sponsored Originator Name")),
            "sponsor_originator_ein_last4": _clean_str(row.get("Sponsored Originator EIN ID (last 4 digits)")),
            "sponsor_originator_nmls_id": _clean_str(row.get("Sponsored Originator NMLS ID")),
            "gift_letter_amount": _clean_num(row.get("Gift Ltr Amt")),
            "gift_letter_source": _clean_str(row.get("Gift Ltr Source")),
            "census_tract": _clean_str(row.get("Census Tract")),
            "underserved_indicator": _clean_str(row.get("Underserved Indicator")),
            "delinquent_reason_code": _clean_str(row.get("Delinquent Reason")),
            "payments_before_first_90_day_delinquent": _clean_int(
                row.get("Payments before First 90 Day Delinquent Reported")
            ),
            "indemnification_flag": _clean_str(row.get("Indem")),

            # ── Enc Data fields for Deep Dive ──
            "underwriter_enc": _clean_str(row.get("Underwriter")),
            "lo_employee_id": _clean_str(row.get("LO Employee ID")),
            "dq_status_enc": _clean_str(row.get("DQ")),
            "hud_reason_code_enc": _clean_str(row.get("HUD Reason Code")),
            "ae_name": _clean_str(row.get("AE Name")),
            "subservicer": _clean_str(row.get("Subservicer")),
            "org_id": _clean_str(row.get("Org ID")),
            "tpo_broker_flag": _clean_str(row.get("TPO Broker")),
            "funded_date": _iso_date(row.get("Fund Date")),
            "closed_date": _iso_date(row.get("Closed Date")),
            # Encompass Data Tab, column BB ("First Pymt Date"). Drives the
            # "Proposed Drop-Off (Next 3 Mo)" column — we use this date as a
            # proxy for HUD's "beginning amortization date" rolling 24-month
            # window. Loans whose First Payment Date is older than
            # (current window start + 3 months) are projected to drop off the
            # office's HUD CR denominator.
            "first_payment_date": _iso_date(
                row.get("First Pymt Date") or row.get("First Payment Date")
            ),
            "lien_position": _clean_str(row.get("Lien Position")),
            "borrower_count": _clean_int(row.get("Borrower Count")),
            "total_income": _clean_num(row.get("Total Income")),
            "is_fthb": str(row.get("FTHB") or "").strip().lower() == "yes",
            "cltv": _clean_num(row.get("CLTV")),
            "interest_rate": _clean_num(row.get("Interest Rate")),
            "insuring_hoc": _clean_str(row.get("Insuring HOC Center")),
        })

    return loans


# Excel "1900-date-system" epoch: serial 1 = 1900-01-01, but Excel mistakenly
# treats 1900 as a leap year, so serial 60 = 1900-02-29 (a date that doesn't
# exist in the proleptic Gregorian calendar). Standard workaround: anchor the
# epoch at 1899-12-30 and add the serial as days. For serials >= 61 this is
# exact; for serials 1..59 there's a 1-day offset, but all the dates we care
# about (First Pymt Date, Fund Date, Closed Date) are post-2020 so the
# correction is irrelevant.
_EXCEL_EPOCH = dt.date(1899, 12, 30)


def _iso_date(v: Any) -> Optional[str]:
    """Coerce a cell value into an ISO `YYYY-MM-DD` date string.

    Accepts ``datetime``/``date`` objects (when openpyxl has already parsed
    the formatted cell), Excel serial numbers (int/float — the common case
    when the cell type is "general" or numeric formatting strips the date
    style), and ISO-shaped strings. Returns ``None`` when the value can't
    be coerced — callers must tolerate missing dates.
    """
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    # Excel serial: numeric or numeric-string. Reject obviously-bogus values
    # (negative; pre-1990; far future) so we never produce a date the rest of
    # the pipeline will then have to ignore.
    if isinstance(v, (int, float)):
        try:
            serial = float(v)
            if 32874 <= serial <= 109573:  # 1990-01-01 .. 2199-12-31
                return (_EXCEL_EPOCH + dt.timedelta(days=int(serial))).isoformat()
        except (ValueError, OverflowError):
            pass
        return None
    s = str(v).strip()
    if not s:
        return None
    # Numeric-string serial (some Excel exports come through as text)
    try:
        serial = float(s)
        if 32874 <= serial <= 109573:
            return (_EXCEL_EPOCH + dt.timedelta(days=int(serial))).isoformat()
    except ValueError:
        pass
    # Already-ISO string — keep the first 10 chars when shaped like a date.
    if len(s) >= 10 and s[4] == '-' and s[7] == '-':
        return s[:10]
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Derived aggregates
# ─────────────────────────────────────────────────────────────────────────────

def _dq_pct(delinquent: int, total: int) -> Optional[float]:
    return round((delinquent / total) * 100, 4) if total > 0 else None


def _bucketize_fico(fico: Optional[int]) -> str:
    if fico is None or fico == 0:
        return "Unknown"
    if fico < 580: return "<580"
    if fico < 620: return "580-619"
    if fico < 660: return "620-659"
    if fico < 680: return "660-679"
    if fico < 700: return "680-699"
    if fico < 740: return "700-739"
    return "740+"


def _bucketize_dti(dti: Optional[float]) -> str:
    if dti is None:
        return "Unknown"
    if dti < 36: return "<36"
    if dti < 43: return "36-42.99"
    if dti < 50: return "43-49.99"
    if dti < 57: return "50-56.99"
    return "57+"


def _bucketize_ltv(ltv: Optional[float]) -> str:
    if ltv is None:
        return "Unknown"
    if ltv < 80: return "<80"
    if ltv < 90: return "80-89.99"
    if ltv < 95: return "90-94.99"
    if ltv < 97: return "95-96.99"
    return "97+"


def _bucketize_risk(n: int) -> str:
    if n >= 5:
        return "5+"
    return str(n)


def build_portfolio_slices(loans: List[dict]) -> List[dict]:
    """Generate flat (dimension, bucket) rows matching fha.portfolio_slices."""
    total = len(loans)
    total_dlq = sum(1 for l in loans if l["is_delinquent"])
    total_retail = [l for l in loans if l["channel"] == "Retail"]
    total_ws = [l for l in loans if l["channel"] == "Wholesale"]
    base_combined = _dq_pct(total_dlq, total)
    base_retail = _dq_pct(sum(1 for l in total_retail if l["is_delinquent"]), len(total_retail))
    base_wholesale = _dq_pct(sum(1 for l in total_ws if l["is_delinquent"]), len(total_ws))

    # Dimension definitions: (dimension_key, getter, bucket_order_map or None)
    def fixed_order(labels):
        return {lbl: i for i, lbl in enumerate(labels)}

    dims: List[Tuple[str, Any, Optional[Dict[str, int]]]] = [
        ("dpa_program", lambda l: l.get("dpa_program") or "Non-DPA", fixed_order(["Boost", "Arrive/Aurora", "Non-DPA"])),
        ("dpa_investor", lambda l: l.get("dpa_investor") or "Unassigned", None),
        ("channel", lambda l: l.get("channel") or "Unknown", fixed_order(["Retail", "Wholesale", "Unknown"])),
        ("fico", lambda l: _bucketize_fico(l.get("fico_score")),
         fixed_order(["<580", "580-619", "620-659", "660-679", "680-699", "700-739", "740+", "Unknown"])),
        ("front_dti", lambda l: _bucketize_dti(l.get("front_dti")),
         fixed_order(["<36", "36-42.99", "43-49.99", "50-56.99", "57+", "Unknown"])),
        ("back_dti", lambda l: _bucketize_dti(l.get("back_dti")),
         fixed_order(["<36", "36-42.99", "43-49.99", "50-56.99", "57+", "Unknown"])),
        ("ltv", lambda l: _bucketize_ltv(l.get("ltv")),
         fixed_order(["<80", "80-89.99", "90-94.99", "95-96.99", "97+", "Unknown"])),
        ("investor", lambda l: l.get("investor_name") or "Unassigned", None),
        ("hud_office", lambda l: l.get("hud_office") or "Unknown", None),
        ("source_of_funds", lambda l: l.get("source_of_funds_group") or l.get("source_of_funds") or "Unknown", None),
        ("employment", lambda l: l.get("employment_type") or "Unknown", None),
        ("aus", lambda l: l.get("aus") or "Unknown", None),
        ("loan_purpose", lambda l: l.get("loan_purpose") or "Unknown", None),
        ("units", lambda l: str(l.get("units") or "Unknown"), None),
        ("risk_indicator_count", lambda l: _bucketize_risk(l.get("risk_indicator_count") or 0),
         fixed_order(["0", "1", "2", "3", "4", "5+"])),
    ]

    rows: List[dict] = []
    for dim_key, getter, order_map in dims:
        agg: Dict[str, Dict[str, int]] = {}
        for l in loans:
            bucket = getter(l) or "Unknown"
            box = agg.setdefault(bucket, {"combined": 0, "retail": 0, "ws": 0,
                                         "cdlq": 0, "rdlq": 0, "wdlq": 0})
            box["combined"] += 1
            if l["is_delinquent"]:
                box["cdlq"] += 1
            if l["channel"] == "Retail":
                box["retail"] += 1
                if l["is_delinquent"]:
                    box["rdlq"] += 1
            elif l["channel"] == "Wholesale":
                box["ws"] += 1
                if l["is_delinquent"]:
                    box["wdlq"] += 1

        bucket_names = list(agg.keys())
        if order_map:
            bucket_names.sort(key=lambda b: (order_map.get(b, 9999), b))
        else:
            bucket_names.sort(key=lambda b: (-agg[b]["combined"], b))

        for order_idx, bucket in enumerate(bucket_names):
            box = agg[bucket]
            comb_pct = _dq_pct(box["cdlq"], box["combined"])
            ret_pct = _dq_pct(box["rdlq"], box["retail"])
            ws_pct = _dq_pct(box["wdlq"], box["ws"])
            rows.append({
                "dimension": dim_key,
                "bucket": bucket,
                "bucket_order": order_idx,
                "combined_population": box["combined"],
                "retail_population": box["retail"],
                "wholesale_population": box["ws"],
                "combined_delinquent": box["cdlq"],
                "retail_delinquent": box["rdlq"],
                "wholesale_delinquent": box["wdlq"],
                "combined_pct": comb_pct,
                "retail_pct": ret_pct,
                "wholesale_pct": ws_pct,
                "baseline_combined": base_combined,
                "baseline_retail": base_retail,
                "baseline_wholesale": base_wholesale,
                "baseline_comparison_combined":
                    round(comb_pct - base_combined, 4) if comb_pct is not None and base_combined is not None else None,
                "baseline_comparison_retail":
                    round(ret_pct - base_retail, 4) if ret_pct is not None and base_retail is not None else None,
                "baseline_comparison_wholesale":
                    round(ws_pct - base_wholesale, 4) if ws_pct is not None and base_wholesale is not None else None,
            })
    return rows


def build_loan_officer_performance(loans: List[dict]) -> List[dict]:
    total_dlq = sum(1 for l in loans if l["is_delinquent"])
    base = _dq_pct(total_dlq, len(loans))

    by_lo: Dict[str, List[dict]] = {}
    for l in loans:
        nmls = l.get("lo_nmls_id") or "unknown"
        by_lo.setdefault(nmls, []).append(l)

    out: List[dict] = []
    for nmls, group in by_lo.items():
        funded = len(group)
        dlq = sum(1 for l in group if l["is_delinquent"])
        pct = _dq_pct(dlq, funded)
        dq_group = [l for l in group if l["is_delinquent"]]

        channels = {l.get("channel") for l in group if l.get("channel")}
        channel = channels.pop() if len(channels) == 1 else None

        out.append({
            "lo_nmls_id": str(nmls),
            "lo_name": _clean_str(group[0].get("loan_officer")),
            "approval_status": None,
            "channel": channel,
            "funded_count": funded,
            "delinquent_count": dlq,
            "delinquency_pct": pct,
            "baseline_comparison": round(pct - base, 4) if pct is not None and base is not None else None,
            "sub_620_count": sum(1 for l in dq_group if l["has_sub_620"]),
            "super_29_dti_count": sum(1 for l in dq_group if l["has_super_29_dti"]),
            "super_50_dti_count": sum(1 for l in dq_group if l["has_super_50_dti"]),
            "super_90_ltv_count": sum(1 for l in dq_group if l["has_super_90_ltv"]),
            "super_95_ltv_count": sum(1 for l in dq_group if l["has_super_95_ltv"]),
            "dpa_count": sum(1 for l in dq_group if l["has_dpa"]),
            "manufactured_count": sum(1 for l in dq_group if l["has_manufactured"]),
            "variable_income_count": sum(1 for l in dq_group if l["has_variable_income"]),
            "super_variable_income_count": sum(1 for l in dq_group if l["has_super_variable_income"]),
            "non_owner_occupied_count": sum(1 for l in dq_group if l["has_non_owner_occupied"]),
            "manual_uw_count": sum(1 for l in dq_group if l["has_manual_uw"]),
            "hud_deficiency_count": sum(1 for l in dq_group if l["has_hud_deficiency"]),
            "gift_grant_count": sum(1 for l in dq_group if l["has_gift_grant"]),
        })
    out.sort(key=lambda r: -(r["delinquency_pct"] or 0))
    return out


def build_risk_indicator_distribution(loans: List[dict]) -> List[dict]:
    counts: Dict[int, Dict[str, int]] = {}
    for l in loans:
        n = min(int(l.get("risk_indicator_count") or 0), 13)
        box = counts.setdefault(n, {"loans": 0, "dlq": 0})
        box["loans"] += 1
        if l["is_delinquent"]:
            box["dlq"] += 1

    base = _dq_pct(sum(c["dlq"] for c in counts.values()), sum(c["loans"] for c in counts.values()))

    out: List[dict] = []
    for n in range(0, 14):
        box = counts.get(n, {"loans": 0, "dlq": 0})
        pct = _dq_pct(box["dlq"], box["loans"])
        out.append({
            "indicator_count": n,
            "loans_count": box["loans"],
            "delinquent_count": box["dlq"],
            "delinquency_pct": pct,
            "baseline_comparison": round(pct - base, 4) if pct is not None and base is not None else None,
        })
    return out


# ─────────────────────────────────────────────────────────────────────────────
# NW Data extension rollups
# ─────────────────────────────────────────────────────────────────────────────

# HUD Neighborhood Watch "Delinquent Reason" code → description.
# Source: HUD Handbook 4000.1 / NW Report legend. Best-effort decoding of
# the code values observed in the NW Data 2 export. Unknown codes fall
# through to "Reason {code}".
DELINQUENT_REASON_CODES: Dict[str, str] = {
    "1":  "Death of principal mortgagor",
    "2":  "Illness of principal mortgagor",
    "3":  "Illness of mortgagor's family member",
    "4":  "Death of mortgagor's family member",
    "5":  "Marital difficulties",
    "6":  "Curtailment of income",
    "7":  "Excessive obligations",
    "8":  "Abandonment of property",
    "9":  "Distant employment transfer",
    "10": "Neighborhood problem",
    "11": "Property problem",
    "12": "Inability to sell property",
    "13": "Inability to rent property",
    "14": "Military service",
    "15": "Other",
    "16": "Unemployment",
    "17": "Business failure",
    "18": "Casualty loss",
    "19": "Energy / environment costs",
    "20": "Servicing problems",
    "21": "Payment adjustment",
    "22": "Payment dispute",
    "23": "Transfer of ownership",
    "24": "Fraud",
    "25": "Incarceration",
}


def _reason_description(code: Optional[str]) -> str:
    if not code:
        return "Not reported"
    s = str(code).strip()
    try:
        s = str(int(float(s)))
    except ValueError:
        pass
    return DELINQUENT_REASON_CODES.get(s, f"Reason {s}")


def build_underwriter_rollup(loans: List[dict]) -> List[dict]:
    """Group loans by underwriter, with SDQ count + credit-rating breakdown.

    NW Data 2 only populates `underwriter_name` for the SDQ population, so
    this rollup naturally limits itself to SDQ-touched underwriters. We skip
    the `Unassigned` bucket (loans that were never in NW Data 2).
    """
    by_uw: Dict[Tuple[str, str], List[dict]] = {}
    for l in loans:
        name = l.get("underwriter_name")
        if not name:
            continue
        uid = l.get("underwriter_id") or ""
        by_uw.setdefault((name, uid), []).append(l)

    total = len(loans)
    total_sdq = sum(1 for l in loans if l.get("is_seriously_delinquent"))
    base = _dq_pct(total_sdq, total) or 0.0

    out: List[dict] = []
    for (name, uid), group in by_uw.items():
        loan_count = len(group)
        sdq_count = sum(1 for l in group if l.get("is_seriously_delinquent"))
        sdq_pct = _dq_pct(sdq_count, loan_count)
        compare_ratio = round((sdq_pct / base) * 100, 2) if sdq_pct is not None and base > 0 else None
        rating_counts: Dict[str, int] = {}
        for l in group:
            rating = (l.get("underwriter_mortgage_credit_rating") or "").strip() or "Unrated"
            rating_counts[rating] = rating_counts.get(rating, 0) + 1
        breakdown = [
            {"rating": k, "count": v}
            for k, v in sorted(rating_counts.items(), key=lambda kv: -kv[1])
        ]
        out.append({
            "underwriter_name": name.strip(),
            "underwriter_id": uid.strip(),
            "loan_count": loan_count,
            "sdq_count": sdq_count,
            "sdq_pct": sdq_pct,
            "compare_ratio": compare_ratio,
            "mortgage_credit_rating_breakdown": breakdown,
        })
    out.sort(key=lambda r: (-r["loan_count"], r["underwriter_name"]))
    return out


def build_delinquency_reason_rollup(loans: List[dict]) -> List[dict]:
    """Group SDQ loans by HUD's Delinquent Reason code."""
    sdq = [l for l in loans if l.get("is_seriously_delinquent")]
    total_sdq = len(sdq)
    counts: Dict[str, int] = {}
    for l in sdq:
        code = (l.get("delinquent_reason_code") or "").strip() or "Not reported"
        try:
            code = str(int(float(code)))
        except ValueError:
            pass
        counts[code] = counts.get(code, 0) + 1

    out: List[dict] = []
    for code, n in sorted(counts.items(), key=lambda kv: -kv[1]):
        pct = round((n / total_sdq) * 100, 2) if total_sdq > 0 else 0.0
        out.append({
            "reason_code": code,
            "reason_description": _reason_description(code),
            "loan_count": n,
            "pct_of_sdq": pct,
        })
    return out


def build_indemnification_loans(loans: List[dict]) -> List[dict]:
    """List loans flagged with an indemnification on the NW export."""
    out: List[dict] = []
    for l in loans:
        flag = (l.get("indemnification_flag") or "").strip()
        if not flag or flag.upper() == "N":
            continue
        out.append({
            "loan_id": l.get("loan_id"),
            "fha_case_number": l.get("fha_case_number"),
            "lo_name": l.get("loan_officer"),
            "indemnification_type": flag,
            "sdq_status": "SDQ" if l.get("is_seriously_delinquent") else "Current",
            "delinquent_status_code": l.get("delinquent_status_code"),
            "months_delinquent": l.get("months_delinquent"),
            "hud_office": l.get("hud_office"),
            "channel": l.get("channel"),
        })
    out.sort(key=lambda r: (r.get("loan_id") or ""))
    return out


def build_sponsor_tpo_detail(loans: List[dict]) -> List[dict]:
    """Per-TPO / sponsored-originator rollup from NW Data sponsor columns.

    NW Data 2 only populates the sponsor columns for the SDQ population, so
    this view is by construction an SDQ-by-TPO breakdown. The compare_ratio
    field is included for symmetry with the underwriter rollup but should
    be interpreted relative to the firm-wide SDQ rate.
    """
    by_tpo: Dict[Tuple[str, str], List[dict]] = {}
    for l in loans:
        name = l.get("sponsor_originator_name")
        if not name:
            continue
        nmls = l.get("sponsor_originator_nmls_id") or ""
        by_tpo.setdefault((name.strip(), str(nmls).strip()), []).append(l)

    total = sum(len(g) for g in by_tpo.values())
    total_sdq = sum(
        1 for g in by_tpo.values() for l in g if l.get("is_seriously_delinquent")
    )
    base = _dq_pct(total_sdq, total) or 0.0

    out: List[dict] = []
    for (name, nmls), group in by_tpo.items():
        loan_count = len(group)
        sdq_count = sum(1 for l in group if l.get("is_seriously_delinquent"))
        sdq_pct = _dq_pct(sdq_count, loan_count)
        compare_ratio = round((sdq_pct / base) * 100, 2) if sdq_pct is not None and base > 0 else None
        sample = group[0]
        out.append({
            "sponsor_originator_name": name,
            "sponsor_originator_nmls_id": nmls or None,
            "sponsor_originator_ein_last4": sample.get("sponsor_originator_ein_last4"),
            "sponsor_id": sample.get("sponsor_id"),
            "loan_count": loan_count,
            "sdq_count": sdq_count,
            "sdq_pct": sdq_pct,
            "compare_ratio": compare_ratio,
        })
    out.sort(key=lambda r: (-r["loan_count"], r["sponsor_originator_name"]))
    return out


# ─────────────────────────────────────────────────────────────────────────────
# HUD Branch → AFN Branch Name bridge (via case number)
# ─────────────────────────────────────────────────────────────────────────────

def _enrich_branch_rows(branch_rows: List[dict], loans: List[dict],
                        nw2_path: Path) -> None:
    """Populate afn_branch_names, hud_offices, afn_org_ids on each
    compare_ratios_branch entry by bridging through NW Data case numbers
    to Encompass loan records."""
    wb = openpyxl.load_workbook(nw2_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_idx = -1
    for i, r in enumerate(rows[:20]):
        if r and r[0] and "ORIGINATING" in str(r[0]).upper():
            header_idx = i
            break
    if header_idx < 0:
        print("  WARNING: could not find NW Data header for branch enrichment")
        return

    # col 0 = Originating ID (FHA Branch ID), col 7 = Case Number
    orig_to_cases: Dict[str, List[str]] = {}
    for r in rows[header_idx + 1:]:
        if not r or r[0] is None:
            continue
        orig_id = str(r[0]).strip()
        case_num = str(r[7]).strip() if r[7] else None
        if orig_id and case_num:
            orig_to_cases.setdefault(orig_id, []).append(case_num)

    case_to_info: Dict[str, dict] = {}
    for loan in loans:
        cn = loan.get("fha_case_number")
        if cn:
            case_to_info[cn] = {
                "branch_name": loan.get("branch_name"),
                "hud_office": loan.get("hud_office"),
                "org_id": loan.get("org_id") or loan.get("branch_nmls_id"),
            }

    enriched = 0
    for br in branch_rows:
        fha_id = br["nmls_id"]
        cases = orig_to_cases.get(fha_id, [])
        if not cases:
            continue

        branch_names = set()
        hud_offices = set()
        org_ids = set()
        for cn in cases:
            info = case_to_info.get(cn)
            if info:
                if info["branch_name"]:
                    branch_names.add(info["branch_name"])
                if info["hud_office"]:
                    hud_offices.add(info["hud_office"])
                if info["org_id"]:
                    org_ids.add(str(info["org_id"]))

        br["afn_branch_names"] = sorted(branch_names) if branch_names else None
        br["hud_offices"] = sorted(hud_offices) if hud_offices else None
        br["afn_org_ids"] = sorted(org_ids) if org_ids else None
        if branch_names:
            if len(branch_names) == 1:
                br["branch_name"] = list(branch_names)[0]
            else:
                br["branch_name"] = f"{len(branch_names)} AFN branches"
            enriched += 1
        if hud_offices:
            if len(hud_offices) == 1:
                br["hud_office"] = list(hud_offices)[0]
            else:
                br["hud_office"] = f"{len(hud_offices)} offices"

    print(f"  Enriched {enriched}/{len(branch_rows)} HUD branches with AFN names")


# ─────────────────────────────────────────────────────────────────────────────
# AI Insights — LLM-generated narrative findings
# ─────────────────────────────────────────────────────────────────────────────

# Allowed Lucide icon names. Keep this list in sync with the AIInsights
# component's `ICON_MAP` so unknown icons can't render as blanks.
_AI_INSIGHT_ICONS = {
    "TrendingUp", "TrendingDown", "AlertTriangle", "Users", "Layers",
    "MapPin", "Building2", "DollarSign", "ShieldAlert", "Activity",
    "BarChart3", "Target", "Flame", "Sparkles",
}
_AI_INSIGHT_TONES = {"red", "yellow", "blue", "green"}

_FALLBACK_AI_INSIGHTS = [
    {
        "icon": "Sparkles",
        "tone": "blue",
        "title": "AI insights unavailable",
        "body": "Live AI analysis could not run for this snapshot. Underlying compare-ratio, delinquency, and DPA data is fully populated — review the panels below for risk findings.",
    },
]


def _round(v: Any, n: int = 1) -> Any:
    try:
        return round(float(v), n)
    except (TypeError, ValueError):
        return None


def _build_ai_facts(snapshot: Dict[str, Any]) -> Dict[str, Any]:
    """Distill the snapshot into a compact facts payload for the LLM.

    We intentionally pre-aggregate so the model never sees raw loan rows
    and the prompt stays well under any context budget.
    """
    facts: Dict[str, Any] = {}
    facts["period"] = snapshot["snapshot_meta"]["label"]

    # ── Company / channel compare ratios ──
    by_scope = {r["scope"]: r for r in snapshot.get("compare_ratios_total", [])}
    facts["compare_ratio_total"] = {
        "company": _round(by_scope.get("total", {}).get("compare_ratio")),
        "retail": _round(by_scope.get("retail", {}).get("compare_ratio")),
        "sponsor_wholesale": _round(by_scope.get("sponsor", {}).get("compare_ratio")),
        "total_loans": by_scope.get("total", {}).get("loans_count"),
        "total_delinquent": by_scope.get("total", {}).get("delinquent_count"),
    }

    # ── HOC roll-up ──
    facts["compare_ratio_by_hoc"] = [
        {
            "hoc": r["hoc_name"],
            "compare_ratio": _round(r.get("compare_ratio")),
            "retail_ratio": _round(r.get("retail_ratio")),
            "sponsor_ratio": _round(r.get("sponsor_ratio")),
            "loans": r.get("loans_count"),
            "delinquent": r.get("delinquent_count"),
        }
        for r in snapshot.get("compare_ratios_hoc", [])
    ]

    # ── Top 10 HUD offices by compare ratio (≥30 loans) ──
    offices = [
        o for o in snapshot.get("compare_ratios_hud_office", [])
        if (o.get("loans_count") or 0) >= 30 and o.get("compare_ratio") is not None
    ]
    offices.sort(key=lambda o: -(o.get("compare_ratio") or 0))
    facts["top_hud_offices_by_compare_ratio"] = [
        {
            "office": o["hud_office"],
            "hoc": o.get("hoc"),
            "compare_ratio": _round(o.get("compare_ratio")),
            "loans": o.get("loans_count"),
            "delinquent": o.get("delinquent_count"),
        }
        for o in offices[:10]
    ]

    # ── DPA program concentration & delinquency ──
    dpa_slices = [
        s for s in snapshot.get("portfolio_slices", [])
        if s.get("dimension") == "dpa_program"
    ]
    facts["dpa_program_concentration"] = [
        {
            "program": s["bucket"],
            "loans": s.get("combined_population"),
            "delinquent": s.get("combined_delinquent"),
            "dq_pct": _round(s.get("combined_pct"), 2),
            "baseline_dq_pct": _round(s.get("baseline_combined"), 2),
            "delta_vs_baseline": _round(s.get("baseline_comparison_combined"), 2),
            "retail_dq_pct": _round(s.get("retail_pct"), 2),
            "wholesale_dq_pct": _round(s.get("wholesale_pct"), 2),
        }
        for s in dpa_slices
        if (s.get("combined_population") or 0) >= 20  # drop noise
    ]

    # ── Channel split ──
    facts["channel_split"] = [
        {
            "channel": s["bucket"],
            "loans": s.get("combined_population"),
            "delinquent": s.get("combined_delinquent"),
            "dq_pct": _round(s.get("combined_pct"), 2),
        }
        for s in snapshot.get("portfolio_slices", [])
        if s.get("dimension") == "channel"
    ]

    # ── FICO bucket distribution (delinquency lens) ──
    facts["fico_buckets"] = [
        {
            "bucket": s["bucket"],
            "loans": s.get("combined_population"),
            "dq_pct": _round(s.get("combined_pct"), 2),
            "delta_vs_baseline": _round(s.get("baseline_comparison_combined"), 2),
        }
        for s in snapshot.get("portfolio_slices", [])
        if s.get("dimension") == "fico" and (s.get("combined_population") or 0) >= 50
    ]

    # ── Top problem LOs (≥30 loans, sorted by delta-vs-baseline) ──
    los = [
        l for l in snapshot.get("loan_officer_performance", [])
        if (l.get("funded_count") or 0) >= 30
        and (l.get("baseline_comparison") or 0) > 0
    ]
    los.sort(key=lambda l: -(l.get("baseline_comparison") or 0))
    facts["top_outlier_loan_officers"] = [
        {
            "name": l.get("lo_name"),
            "channel": l.get("channel"),
            "funded": l.get("funded_count"),
            "delinquent": l.get("delinquent_count"),
            "dq_pct": _round(l.get("delinquency_pct"), 2),
            "delta_vs_baseline": _round(l.get("baseline_comparison"), 2),
        }
        for l in los[:8]
    ]

    # ── Risk-indicator distribution (count of stacked indicators per loan) ──
    facts["risk_indicator_distribution"] = [
        {
            "indicator_count": r["indicator_count"],
            "loans": r.get("loans_count"),
            "dq_pct": _round(r.get("delinquency_pct"), 2),
        }
        for r in snapshot.get("risk_indicator_distribution", [])
        if (r.get("loans_count") or 0) > 0
    ]

    # ── Projections ─ forward-looking risk at 1/3/6mo × best/base/worst ──
    # Pre-aggregated so the LLM can flag offices projected to cross 150/200
    # without seeing raw loan rows.
    proj = snapshot.get("projections") or {}
    if proj:
        facts["projection_assumptions"] = proj.get("assumptions")
        # Top 8 offices projected to be in breach/watch at 3mo base.
        proj_offices = proj.get("offices") or []
        top_3mo_base = sorted(
            [
                o for o in proj_offices
                if o["horizons"]["3mo"]["scenarios"]["base"]["projected_threshold_status"]
                in ("breach", "watch")
            ],
            key=lambda o: -(
                o["horizons"]["3mo"]["scenarios"]["base"]["projected_compare_ratio"] or 0
            ),
        )[:8]
        facts["projections_top_offices_3mo_base"] = [
            {
                "office": o["office_name"],
                "hoc": o["hoc"],
                "loan_count": o["loan_count_current"],
                "current_compare_ratio": o["current_compare_ratio"],
                "current_status": o["current_threshold_status"],
                "projected_dropoffs_3mo": o["horizons"]["3mo"]["projected_dropoffs"],
                "projected_compare_ratio_3mo_base": o["horizons"]["3mo"]["scenarios"]["base"]["projected_compare_ratio"],
                "projected_compare_ratio_3mo_worst": o["horizons"]["3mo"]["scenarios"]["worst"]["projected_compare_ratio"],
                "projected_compare_ratio_3mo_best": o["horizons"]["3mo"]["scenarios"]["best"]["projected_compare_ratio"],
                "projected_status_3mo_base": o["horizons"]["3mo"]["scenarios"]["base"]["projected_threshold_status"],
                "projected_status_3mo_worst": o["horizons"]["3mo"]["scenarios"]["worst"]["projected_threshold_status"],
            }
            for o in top_3mo_base
        ]
        # Offices currently safe that cross into watch/breach at any horizon/scenario.
        crossings_flat: List[Dict[str, Any]] = []
        for o in proj_offices:
            for c in o.get("threshold_crossings") or []:
                crossings_flat.append({
                    "office": o["office_name"],
                    "hoc": o["hoc"],
                    "loan_count": o["loan_count_current"],
                    "horizon_months": c["horizon_months"],
                    "scenario": c["scenario"],
                    "from_status": c["from_status"],
                    "to_status": c["to_status"],
                    "current_compare_ratio": c["current_compare_ratio"],
                    "projected_compare_ratio": c["projected_compare_ratio"],
                })
        # Prioritize: safe→breach, then safe→watch, then watch→breach; by CR desc.
        def _cross_prio(c: Dict[str, Any]) -> Tuple[int, float]:
            pri = {("safe", "breach"): 0, ("safe", "watch"): 1, ("watch", "breach"): 2}
            return (pri.get((c["from_status"], c["to_status"]), 9),
                    -(c["projected_compare_ratio"] or 0))
        crossings_flat.sort(key=_cross_prio)
        facts["projections_threshold_crossings"] = crossings_flat[:12]
        # HOC-level roll-up for orientation.
        facts["projections_by_hoc_3mo"] = [
            {
                "hoc": name,
                "current_compare_ratio": block.get("current_compare_ratio"),
                "projected_compare_ratio_base": block["horizons"]["3mo"]["scenarios"]["base"]["projected_compare_ratio"],
                "projected_compare_ratio_worst": block["horizons"]["3mo"]["scenarios"]["worst"]["projected_compare_ratio"],
                "projected_compare_ratio_best": block["horizons"]["3mo"]["scenarios"]["best"]["projected_compare_ratio"],
                "projected_dropoffs": block["horizons"]["3mo"]["projected_dropoffs"],
            }
            for name, block in (proj.get("hocs") or {}).items()
        ]

    return facts


def _normalize_ai_insight(item: Any) -> Optional[Dict[str, Any]]:
    """Validate one model-returned insight, dropping junk silently.

    Backwards-compatible with the pre-projections schema (icon/tone/title/body).
    When present, the extended projection fields are validated and included;
    unknown extras are dropped so the UI can rely on a stable shape.
    """
    if not isinstance(item, dict):
        return None
    icon = str(item.get("icon") or "").strip()
    tone = str(item.get("tone") or "").strip().lower()
    title = str(item.get("title") or "").strip()
    body = str(item.get("body") or "").strip()
    if icon not in _AI_INSIGHT_ICONS:
        icon = "Sparkles"
    if tone not in _AI_INSIGHT_TONES:
        tone = "blue"
    if not title or not body:
        return None
    out: Dict[str, Any] = {"icon": icon, "tone": tone, "title": title, "body": body}

    # ── Extended projection fields (all optional) ──
    def _opt_num(v: Any) -> Optional[float]:
        try:
            return None if v is None else float(v)
        except (TypeError, ValueError):
            return None

    def _opt_int(v: Any) -> Optional[int]:
        try:
            return None if v is None else int(v)
        except (TypeError, ValueError):
            return None

    projected_ratio = _opt_num(item.get("projected_ratio"))
    horizon_months = _opt_int(item.get("horizon_months"))
    if horizon_months not in (1, 3, 6, None):
        horizon_months = None
    scenario = str(item.get("scenario") or "").strip().lower() or None
    if scenario not in ("best", "base", "worst", None):
        scenario = None
    crosses_threshold_val = item.get("crosses_threshold")
    if crosses_threshold_val is None:
        crosses_threshold: Optional[int] = None
    else:
        # Accept either an integer (150/200) or a bool (rare from the LLM).
        try:
            cti = int(crosses_threshold_val)
            crosses_threshold = cti if cti in (150, 200) else None
        except (TypeError, ValueError):
            crosses_threshold = None
    confidence = str(item.get("confidence") or "").strip().lower() or None
    if confidence not in ("low", "medium", "high", None):
        confidence = None

    # Only attach when at least one extended field is non-null — keeps the
    # legacy shape tidy for non-projection insights.
    if any(
        v is not None
        for v in (projected_ratio, horizon_months, scenario, crosses_threshold, confidence)
    ):
        out["projected_ratio"] = projected_ratio
        out["horizon_months"] = horizon_months
        out["scenario"] = scenario
        out["crosses_threshold"] = crosses_threshold
        out["confidence"] = confidence
    return out


def _build_ai_prompt(snapshot: Dict[str, Any]) -> tuple[str, str]:
    """Assemble the (system_prompt, user_prompt) pair used for AI insights.

    Extracted so the Azure OpenAI and LiteLLM code paths share identical
    prompt semantics — the model/provider is the only thing that varies.
    """
    facts = _build_ai_facts(snapshot)
    has_projections = bool(snapshot.get("projections"))
    projection_guidance = (
        " You are also given forward-looking projections at 1/3/6-month horizons "
        "under best/base/worst scenarios (the \u00b110% delinquency lever). "
        "AT LEAST ONE of the 4 insights MUST be projection-based: specifically, "
        "flag any office that is currently safe (< 150) but projects to cross "
        "150 or 200 under any scenario within 3 or 6 months, or any currently-watch "
        "office (150\u2013199) projecting to breach 200. Cite the projected compare "
        "ratio, horizon, and scenario by name in the body. When your insight is "
        "projection-based, ALSO populate these extra JSON fields on the insight: "
        "`projected_ratio` (number), `horizon_months` (1|3|6), "
        "`scenario` ('best'|'base'|'worst'), "
        "`crosses_threshold` (150 or 200, whichever is crossed), and "
        "`confidence` ('low'|'medium'|'high' \u2014 use 'high' for offices with >=100 "
        "loans and consistent base/worst signal, 'medium' otherwise, 'low' for "
        "offices with fewer than 30 loans). Leave those fields off for non-projection "
        "insights."
        if has_projections
        else ""
    )
    system_prompt = (
        "You are a credit-risk analyst writing for AFN's FHA Risk Committee. "
        "You will be given a single month's pre-aggregated FHA portfolio facts "
        "(HUD compare ratios, delinquency rates, DPA concentrations, channel "
        "mix, HUD field-office hotspots, and outlier loan officers). "
        "Generate EXACTLY 4 distinct, non-overlapping insights that a risk "
        "manager would actually surface in the next committee meeting. "
        "Each insight must reference at least one concrete number from the "
        "facts (compare ratio, dq %, loan count, etc.). Avoid generic "
        "statements; call out specific HOCs, HUD offices, or DPA programs. "
        "Prioritize: (1) compare-ratio outliers, (2) delinquency anomalies, "
        "(3) DPA program concentration / drift, (4) channel or geographic "
        "concentration."
        + projection_guidance
        + " Return strict JSON of the form: "
        '{"insights": [{"icon": <one of: ' + ", ".join(sorted(_AI_INSIGHT_ICONS)) + '>, '
        '"tone": <one of: red|yellow|green|blue>, '
        '"title": <short headline, ≤ 80 chars>, '
        '"body": <2 sentences max, ≤ 280 chars>}]}.\n\n'
        "Tone guidance: 'red' = material risk needing action, 'yellow' = "
        "watch-item / drift, 'blue' = neutral structural observation, "
        "'green' = positive / improving trend."
    )
    user_prompt = json.dumps(facts, indent=2, default=str)
    return system_prompt, user_prompt


def _parse_ai_response(raw: str) -> List[Dict[str, Any]]:
    """Parse and validate an LLM JSON response into normalized insights.

    Returns a copy of ``_FALLBACK_AI_INSIGHTS`` on any parse/validation
    failure so callers never see an empty result.
    """
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        # Strip code fences if the model wrapped the JSON
        cleaned = raw.strip().strip("`")
        if cleaned.lower().startswith("json"):
            cleaned = cleaned[4:].strip()
        try:
            parsed = json.loads(cleaned)
        except json.JSONDecodeError:
            print("  WARN: LLM returned non-JSON — falling back")
            return list(_FALLBACK_AI_INSIGHTS)

    items = parsed.get("insights") if isinstance(parsed, dict) else parsed
    if not isinstance(items, list):
        print("  WARN: LLM JSON missing 'insights' list — falling back")
        return list(_FALLBACK_AI_INSIGHTS)

    insights = [n for n in (_normalize_ai_insight(i) for i in items) if n is not None]
    if not insights:
        print("  WARN: LLM returned 0 valid insights — falling back")
        return list(_FALLBACK_AI_INSIGHTS)

    # Cap at 4. Pad with the fallback shape if model under-delivered.
    insights = insights[:4]
    if len(insights) < 4:
        print(f"  NOTE: LLM returned {len(insights)} insights (expected 4)")
    return insights


def _build_ai_insights_azure(
    snapshot: Dict[str, Any],
    *,
    endpoint: str,
    deployment: str,
    api_key: str,
    api_version: str,
) -> List[Dict[str, Any]]:
    """Call Azure OpenAI (chat.completions) for AI insights.

    Uses the ``AzureOpenAI`` client from the ``openai`` package. On any
    error, returns the canned fallback so snapshot build never breaks.
    """
    try:
        from openai import AzureOpenAI  # type: ignore
    except ImportError:
        print("  WARN: openai package not installed / AzureOpenAI unavailable — skipping AI insights")
        return list(_FALLBACK_AI_INSIGHTS)

    system_prompt, user_prompt = _build_ai_prompt(snapshot)

    try:
        client = AzureOpenAI(
            api_version=api_version,
            azure_endpoint=endpoint,
            api_key=api_key,
            timeout=60.0,
        )
        print(f"  Calling Azure OpenAI ({endpoint}, deployment={deployment})…")
        resp = client.chat.completions.create(
            model=deployment,  # NB: Azure uses the deployment name here
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            temperature=0.2,
            response_format={"type": "json_object"},
            max_tokens=1200,
        )
        raw = resp.choices[0].message.content or ""
    except Exception as e:
        print(f"  WARN: AI insight call failed ({e!r}) — falling back")
        return list(_FALLBACK_AI_INSIGHTS)

    return _parse_ai_response(raw)


def _build_ai_insights_litellm(
    snapshot: Dict[str, Any],
    *,
    base_url: str,
    api_key: str,
    model: str,
) -> List[Dict[str, Any]]:
    """Call the legacy AFN LiteLLM proxy for AI insights.

    Kept as a backward-compatible fallback for local dev environments that
    have LiteLLM configured but not Azure OpenAI. In prod (Azure Container
    Apps) this path is unreachable because the tailnet IP is not routable.
    """
    try:
        from openai import OpenAI  # type: ignore
    except ImportError:
        print("  WARN: openai package not installed — skipping AI insights")
        return list(_FALLBACK_AI_INSIGHTS)

    system_prompt, user_prompt = _build_ai_prompt(snapshot)

    try:
        client = OpenAI(base_url=base_url, api_key=api_key, timeout=60.0)
        print(f"  Calling LiteLLM proxy ({base_url}, model={model})…")
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            temperature=0.2,
            response_format={"type": "json_object"},
            max_tokens=1200,
        )
        raw = resp.choices[0].message.content or ""
    except Exception as e:
        print(f"  WARN: AI insight call failed ({e!r}) — falling back")
        return list(_FALLBACK_AI_INSIGHTS)

    return _parse_ai_response(raw)


# ─────────────────────────────────────────────────────────────────────────────
# Portfolio Risk Factor bullets — baked at build time (PR A of 2)
#
# The FHA dashboard's Executive Summary card renders 6 AI-generated bullets
# summarizing risk-factor trends. Historically the frontend called
# /api/ai-analysis on every period load. Baking the bullets into the
# snapshot removes that per-load LLM cost and gives every reviewer the same
# committee-grade summary.
#
# The prompt is loaded from data/prompts/risk-factor-analysis.system.md so
# the future /api/ai-analysis "regenerate" endpoint (PR B) can reuse the same
# file byte-for-byte. Do NOT inline-tune the prompt here.
# ─────────────────────────────────────────────────────────────────────────────

RISK_FACTOR_PROMPT_PATH = REPO_ROOT / "data" / "prompts" / "risk-factor-analysis.system.md"
RISK_FACTOR_BULLETS_SCHEMA_VERSION = 1
_VALID_RISK_SEVERITIES = {"red", "yellow", "green", "neutral"}


def _load_risk_factor_prompt() -> Optional[str]:
    """Load the shared risk-factor system prompt.

    Returns the prompt body with any leading HTML comment (used for
    developer notes at the top of the file) stripped, or ``None`` if the
    file is missing or unreadable. Callers treat ``None`` as a fatal-for-
    this-feature signal and return an empty bullet list.
    """
    try:
        raw = RISK_FACTOR_PROMPT_PATH.read_text(encoding="utf-8")
    except FileNotFoundError:
        print(
            "  WARN: risk-factor prompt missing at "
            f"{RISK_FACTOR_PROMPT_PATH} — skipping risk_factor_bullets"
        )
        return None
    except OSError as e:  # pragma: no cover — filesystem edge
        print(f"  WARN: could not read {RISK_FACTOR_PROMPT_PATH} ({e!r}) — skipping risk_factor_bullets")
        return None
    # Strip a single leading HTML comment (developer notes) if present.
    import re as _re
    stripped = _re.sub(r"^<!--.*?-->\s*", "", raw, count=1, flags=_re.DOTALL)
    return stripped


def _rf_group_by(
    loans: List[dict],
    getter,
    *,
    min_count: int = 20,
) -> List[Dict[str, Any]]:
    """Port of ``groupByField`` from ``src/lib/aiAnalysis.ts``.

    Groups loans by a string label, drops buckets below ``min_count`` (the TS
    dashboard behavior), and returns rows sorted by DQ rate descending.
    """
    groups: Dict[str, Dict[str, int]] = {}
    for l in loans:
        raw = getter(l)
        key = str(raw) if raw not in (None, "", 0) or raw == 0 else "Unknown"
        # Normalize "" / None to "Unknown" to match TS: `key = getter(l) || 'Unknown'`.
        if not key or key == "None":
            key = "Unknown"
        g = groups.setdefault(key, {"total": 0, "dlq": 0})
        g["total"] += 1
        if l.get("is_delinquent"):
            g["dlq"] += 1
    rows = [
        {
            "label": label,
            "total": v["total"],
            "dlq": v["dlq"],
            "dqRate": (v["dlq"] / v["total"] * 100.0) if v["total"] > 0 else 0.0,
        }
        for label, v in groups.items()
        if v["total"] >= min_count
    ]
    rows.sort(key=lambda r: -r["dqRate"])
    return rows


def _rf_risk_indicator_label(cnt: Optional[int]) -> str:
    """Match TS: ``cnt >= 5 ? '5+' : String(cnt)``."""
    n = cnt if isinstance(cnt, int) else 0
    return "5+" if n >= 5 else str(n)


def _rf_channel_summary(loans: List[dict], channel: str) -> Dict[str, Any]:
    ch = [l for l in loans if l.get("channel") == channel]
    total = len(ch)
    dpa = [l for l in ch if l.get("has_dpa")]
    dpa_dq = sum(1 for l in dpa if l.get("is_delinquent"))
    dlq = sum(1 for l in ch if l.get("is_delinquent"))
    return {
        "totalLoans": total,
        "dpaConc": (len(dpa) / total * 100.0) if total > 0 else 0.0,
        "overallDQRate": (dlq / total * 100.0) if total > 0 else 0.0,
        "dpaDQRate": (dpa_dq / len(dpa) * 100.0) if dpa else 0.0,
    }


def _rf_fico_buckets(loans: List[dict]) -> List[Dict[str, Any]]:
    """Port of ``computeFICO`` from ``src/lib/computeData.ts``."""
    buckets = [
        ("<580", 0, 579),
        ("580-619", 580, 619),
        ("620-659", 620, 659),
        ("660-679", 660, 679),
        ("680-699", 680, 699),
        ("700-739", 700, 739),
        ("740+", 740, 999),
    ]
    out: List[Dict[str, Any]] = []
    for label, lo, hi in buckets:
        in_bucket = [l for l in loans if (l.get("fico_score") or 0) >= lo and (l.get("fico_score") or 0) <= hi]
        standard = [l for l in in_bucket if l.get("program_type") == "Standard"]
        dpa = [l for l in in_bucket if l.get("program_type") == "DPA"]
        s_dq = sum(1 for l in standard if l.get("is_delinquent"))
        d_dq = sum(1 for l in dpa if l.get("is_delinquent"))
        out.append({
            "label": label,
            "standardDQ": (s_dq / len(standard) * 100.0) if standard else 0.0,
            "dpaDQ": (d_dq / len(dpa) * 100.0) if dpa else 0.0,
            "dpaTotal": len(dpa),
        })
    return out


def _rf_offices(snapshot: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Approximate the TS OfficeSummary shape from HUD Office + loan rows.

    Only the fields consumed by ``buildDataSummary`` in aiAnalysis.ts are
    populated. This intentionally does NOT re-run the full Revised-CR /
    Enhanced Guidelines pipeline that ``computeOffices`` performs — the
    Portfolio Risk Factor bullets don't need it, and duplicating that math
    in Python would create a second source of truth for committee numbers.
    """
    hud_rows = {(r.get("hud_office") or "").upper().strip(): r for r in snapshot.get("compare_ratios_hud_office", [])}
    by_office: Dict[str, List[dict]] = {}
    for l in snapshot.get("loans", []):
        by_office.setdefault(l.get("hud_office") or "", []).append(l)

    offices: List[Dict[str, Any]] = []
    for name, o_loans in by_office.items():
        total = len(o_loans)
        retail = [l for l in o_loans if l.get("channel") == "Retail"]
        ws = [l for l in o_loans if l.get("channel") == "Wholesale"]
        total_dlq = sum(1 for l in o_loans if l.get("is_delinquent"))
        hud = hud_rows.get(name.upper().strip())
        offices.append({
            "name": name,
            "totalCR": (hud.get("compare_ratio") if hud else None) or 0,
            "retailCR": (hud.get("retail_ratio") if hud else None),
            "wsCR": (hud.get("sponsor_ratio") if hud else None),
            "totalLoans": total,
            "totalDLQ": total_dlq,
            "retailBoostDLQ": sum(1 for l in retail if l.get("is_delinquent") and l.get("is_boost")),
            "wsBoostDLQ": sum(1 for l in ws if l.get("is_delinquent") and l.get("is_boost")),
            "retailOtherDPADLQ": sum(1 for l in retail if l.get("is_delinquent") and l.get("has_dpa") and not l.get("is_boost")),
            "wsOtherDPADLQ": sum(1 for l in ws if l.get("is_delinquent") and l.get("has_dpa") and not l.get("is_boost")),
            "retailNonDPADLQ": sum(1 for l in retail if l.get("is_delinquent") and not l.get("has_dpa")),
            "wsNonDPADLQ": sum(1 for l in ws if l.get("is_delinquent") and not l.get("has_dpa")),
            # `revisedTotal*CR` are Revised-CR outputs (Enhanced Guidelines math)
            # not reproduced in Python. Surfaced as ``None`` — the prompt template
            # renders them as "None" which is honest for the build-time bake.
            "revisedTotalCR": None,
            "revisedRetailCR": None,
            "revisedWSCR": None,
            "retailDPAConc": (len([l for l in retail if l.get("has_dpa")]) / len(retail) * 100.0) if retail else 0.0,
            "wsDPAConc": (len([l for l in ws if l.get("has_dpa")]) / len(ws) * 100.0) if ws else 0.0,
            "totalDPAConc": (len([l for l in o_loans if l.get("has_dpa")]) / total * 100.0) if total else 0.0,
        })
    return offices


def _rf_dpa_programs(loans: List[dict]) -> List[Dict[str, Any]]:
    """Port of ``computeDPAPrograms`` (program × investor rollup)."""
    def _prog_label(raw: Any) -> str:
        s = (str(raw) if raw is not None else "").strip()
        if not s:
            return "Non-DPA"
        lower = s.lower()
        if "boost" in lower:
            return "Boost"
        if "arrive" in lower or "aurora" in lower:
            return "Arrive/Aurora"
        return s

    def _inv_label(raw: Any) -> str:
        s = (str(raw) if raw is not None else "").strip()
        return s or "Unassigned"

    dpa_loans = [l for l in loans if l.get("has_dpa")]
    total_dpa = len(dpa_loans)

    by_program: Dict[str, List[dict]] = {}
    for l in dpa_loans:
        by_program.setdefault(_prog_label(l.get("dpa_program")), []).append(l)

    programs: List[Dict[str, Any]] = []
    for program, p_loans in by_program.items():
        p_total = len(p_loans)
        p_dlq = sum(1 for l in p_loans if l.get("is_delinquent"))
        by_investor: Dict[str, List[dict]] = {}
        for l in p_loans:
            # Committee-relevant investor is `investor_name` (matches TS), not `dpa_investor`.
            by_investor.setdefault(_inv_label(l.get("investor_name")), []).append(l)
        investors = [
            {
                "investor": inv,
                "totalLoans": len(i_loans),
                "delinquent": sum(1 for l in i_loans if l.get("is_delinquent")),
                "dqRate": (sum(1 for l in i_loans if l.get("is_delinquent")) / len(i_loans) * 100.0) if i_loans else 0.0,
                "pctOfProgramVolume": (len(i_loans) / p_total * 100.0) if p_total else 0.0,
                "pctOfDPAVolume": (len(i_loans) / total_dpa * 100.0) if total_dpa else 0.0,
            }
            for inv, i_loans in by_investor.items()
        ]
        investors.sort(key=lambda i: -i["delinquent"])
        programs.append({
            "program": program,
            "totalLoans": p_total,
            "delinquent": p_dlq,
            "dqRate": (p_dlq / p_total * 100.0) if p_total else 0.0,
            "pctOfDPAVolume": (p_total / total_dpa * 100.0) if total_dpa else 0.0,
            "investors": investors,
        })
    programs.sort(key=lambda p: -p["delinquent"])
    return programs


def _build_risk_factor_facts(snapshot: Dict[str, Any]) -> str:
    """Port of ``buildDataSummary`` from ``src/lib/aiAnalysis.ts``.

    Returns the exact same free-form string the frontend today assembles
    from ``DashboardData`` so the LLM sees identical facts whether the
    bullets are baked (this function) or regenerated live via the SWA
    proxy (PR B).

    Any field the TS version references but that isn't populated on the
    snapshot is defaulted to a safe zero / "None" so the resulting prompt
    text is stable and readable — the LLM will just skip missing dimensions.
    """
    loans = snapshot.get("loans", []) or []
    total_loans = len(loans)
    total_dlq = sum(1 for l in loans if l.get("is_delinquent"))
    total_dpa = sum(1 for l in loans if l.get("has_dpa"))
    overall_dq_rate = (total_dlq / total_loans * 100.0) if total_loans else 0.0
    dpa_portfolio_conc = (total_dpa / total_loans * 100.0) if total_loans else 0.0

    # ── Program composition (Standard vs DPA DQ%) ──
    standard_loans = [l for l in loans if l.get("program_type") == "Standard"]
    dpa_loans = [l for l in loans if l.get("program_type") == "DPA"]
    standard_dq = (sum(1 for l in standard_loans if l.get("is_delinquent")) / len(standard_loans) * 100.0) if standard_loans else 0.0
    dpa_dq = (sum(1 for l in dpa_loans if l.get("is_delinquent")) / len(dpa_loans) * 100.0) if dpa_loans else 0.0
    multiplier = f"{dpa_dq / standard_dq:.1f}" if standard_dq > 0 else "N/A"

    retail_summary = _rf_channel_summary(loans, "Retail")
    ws_summary = _rf_channel_summary(loans, "Wholesale")

    offices = _rf_offices(snapshot)
    term_offices = [o for o in offices if o["totalCR"] > 200 and o["totalLoans"] > 100]
    cw_offices = [
        o for o in offices if
        (o["totalCR"] > 150 and o["totalCR"] <= 200 and o["totalLoans"] >= 100)
        or (o["totalCR"] > 200 and o["totalLoans"] < 100)
    ]

    def _fmt_num_or_na(v):
        return "N/A" if v is None else v

    term_offices.sort(key=lambda o: -o["totalCR"])
    term_details_lines: List[str] = []
    for o in term_offices:
        boost_dlq = o["retailBoostDLQ"] + o["wsBoostDLQ"]
        other_dpa_dlq = o["retailOtherDPADLQ"] + o["wsOtherDPADLQ"]
        non_dpa_dlq = o["retailNonDPADLQ"] + o["wsNonDPADLQ"]
        term_details_lines.append(
            f"  - {o['name']}: Total CR {o['totalCR']}% "
            f"(Retail {_fmt_num_or_na(o['retailCR'])}%, WS {_fmt_num_or_na(o['wsCR'])}%), "
            f"{o['totalLoans']} loans, {o['totalDLQ']} DLQ "
            f"({non_dpa_dlq} Non-DPA, {boost_dlq} Boost, {other_dpa_dlq} Other DPA), "
            f"Revised CR after Boost removal: {_fmt_num_or_na(o['revisedTotalCR'])}% "
            f"(Retail {_fmt_num_or_na(o['revisedRetailCR'])}%, WS {_fmt_num_or_na(o['revisedWSCR'])}%), "
            f"DPA Conc: Retail {o['retailDPAConc']:.1f}% / WS {o['wsDPAConc']:.1f}%"
        )
    term_details = "\n".join(term_details_lines)

    cw_offices.sort(key=lambda o: -o["totalCR"])
    cw_top5_lines: List[str] = [
        f"  - {o['name']}: Total CR {o['totalCR']}%, {o['totalLoans']} loans, {o['totalDLQ']} DLQ, DPA Conc: {o['totalDPAConc']:.1f}%"
        for o in cw_offices[:5]
    ]
    cw_top5 = "\n".join(cw_top5_lines)

    programs = _rf_dpa_programs(loans)
    programs.sort(key=lambda p: -p["delinquent"])
    program_breakdown_lines: List[str] = []
    for p in programs:
        investor_lines = [
            f"      · Investor {i['investor']}: {i['totalLoans']} loans, {i['delinquent']} DLQ ({i['dqRate']:.1f}%), {i['pctOfProgramVolume']:.1f}% of program"
            for i in p["investors"][:5]
        ]
        prog_line = (
            f"  - Program {p['program']}: {p['totalLoans']} loans, {p['delinquent']} DLQ "
            f"({p['dqRate']:.1f}%), {p['pctOfDPAVolume']:.1f}% of DPA volume"
        )
        if investor_lines:
            prog_line += "\n" + "\n".join(investor_lines)
        program_breakdown_lines.append(prog_line)
    program_breakdown = "\n".join(program_breakdown_lines)

    fico_buckets = _rf_fico_buckets(loans)
    fico_lines = "\n".join(
        f"  {b['label']}: Standard {b['standardDQ']:.1f}%, DPA {b['dpaDQ']:.1f}% ({b['dpaTotal']} DPA loans)"
        for b in fico_buckets
    )

    # ── Trend dimensions (raw group labels from Encompass) ──
    aus_types = _rf_group_by(loans, lambda l: l.get("aus"), min_count=10)
    manual = [l for l in loans if "MANUAL" in (l.get("aus") or "").upper()]
    auto = [l for l in loans if "MANUAL" not in (l.get("aus") or "").upper() and (l.get("aus") or "") != ""]
    manual_uw_rate = (len(manual) / total_loans * 100.0) if total_loans else 0.0
    manual_uw_dq = (sum(1 for l in manual if l.get("is_delinquent")) / len(manual) * 100.0) if manual else 0.0
    auto_uw_dq = (sum(1 for l in auto if l.get("is_delinquent")) / len(auto) * 100.0) if auto else 0.0

    ltv_groups = _rf_group_by(loans, lambda l: l.get("ltv_group"))
    fthb_groups = _rf_group_by(loans, lambda l: l.get("fthb"), min_count=10)
    dti_groups = _rf_group_by(loans, lambda l: l.get("dti_back_end_group"))
    pay_shock_groups = _rf_group_by(loans, lambda l: l.get("payment_shock_group"))
    sof_groups = _rf_group_by(loans, lambda l: l.get("source_of_funds_group"))
    reserves_groups = _rf_group_by(loans, lambda l: l.get("reserves_group"))
    risk_ind_groups = _rf_group_by(
        loans, lambda l: _rf_risk_indicator_label(l.get("risk_indicator_count")), min_count=10
    )
    gg_groups = _rf_group_by(loans, lambda l: l.get("gift_grant_group"))

    def _fmt_trend(rows: List[Dict[str, Any]], label_prefix: str = "", label_suffix: str = "") -> str:
        return "\n".join(
            f"  {label_prefix}{r['label']}{label_suffix}: {r['dqRate']:.1f}% ({r['dlq']}/{r['total']})"
            for r in rows
        )

    return f"""FHA LOAN PORTFOLIO ANALYSIS DATA:

PORTFOLIO OVERVIEW:
- Total Loans: {total_loans:,}
- Overall DQ Rate: {overall_dq_rate:.2f}%
- DPA Portfolio Concentration: {dpa_portfolio_conc:.1f}%
- Program DQ Rates: Standard FHA {standard_dq:.2f}%, DPA {dpa_dq:.2f}% ({multiplier}x standard rate)
- NOTE: "FUEL" is not a distinct program — it was Standard FHA run through the Wholesale channel. Use the Retail vs Wholesale channel breakdown below to see what was previously labeled "FUEL" performance (wholesale-channel Standard FHA).

CHANNEL COMPARISON:
- Retail: {retail_summary['totalLoans']} loans, DPA Conc {retail_summary['dpaConc']:.1f}%, DQ Rate {retail_summary['overallDQRate']:.2f}%, DPA DQ {retail_summary['dpaDQRate']:.2f}%
- Wholesale: {ws_summary['totalLoans']} loans, DPA Conc {ws_summary['dpaConc']:.1f}%, DQ Rate {ws_summary['overallDQRate']:.2f}%, DPA DQ {ws_summary['dpaDQRate']:.2f}%

TERMINATION RISK OFFICES ({len(term_offices)} offices, >200% CR + >100 loans):
{term_details or '  None'}

TOP 5 CREDIT WATCH OFFICES:
{cw_top5 or '  None'}
Total Credit Watch: {len(cw_offices)} offices

DPA PROGRAM × INVESTOR BREAKDOWN (primary = DPA Program, secondary = DPA Investor):
{program_breakdown}

FICO ANALYSIS:
{fico_lines}

UNDERWRITING & RISK FACTOR TRENDS:

AUS Type DQ Rates:
{_fmt_trend(aus_types)}
  Manual UW = {manual_uw_rate:.1f}% of portfolio, DQ rate {manual_uw_dq:.1f}% vs Auto {auto_uw_dq:.1f}%

LTV Group DQ Rates (higher LTV = more risk):
{_fmt_trend(ltv_groups)}

First-Time Homebuyer DQ Rates:
{_fmt_trend(fthb_groups, label_prefix='FTHB=')}

DTI Back-End Group DQ Rates:
{_fmt_trend(dti_groups)}

Payment Shock Group DQ Rates:
{_fmt_trend(pay_shock_groups)}

Source of Funds DQ Rates:
{_fmt_trend(sof_groups)}

Reserves (months) DQ Rates:
{_fmt_trend(reserves_groups, label_suffix=' months')}

Risk Indicator Count DQ Rates (layered risk):
{_fmt_trend(risk_ind_groups, label_suffix=' indicators')}

Gift/Grant Funding % DQ Rates:
{_fmt_trend(gg_groups)}

KEY THRESHOLDS:
- Compare Ratio >200%: Termination risk (HUD can suspend underwriting)
- Compare Ratio 150-200%: Credit watch
- DPA Concentration >40%: High risk
- Each HUD office can independently enforce at >200%"""


def _normalize_risk_factor_bullet(item: Any) -> Optional[Dict[str, str]]:
    """Validate one model-returned bullet — drop junk silently."""
    if not isinstance(item, dict):
        return None
    text = str(item.get("text") or "").strip()
    severity = str(item.get("severity") or "").strip().lower()
    if severity not in _VALID_RISK_SEVERITIES:
        severity = "neutral"
    if not text:
        return None
    return {"text": text, "severity": severity}


def _parse_risk_factor_response(raw: str) -> List[Dict[str, str]]:
    """Parse the LLM JSON body, extracting only ``executiveSummary`` bullets.

    Returns ``[]`` on any parse/validation failure — never raises. Action
    items are intentionally discarded; PR B introduces the write-back path
    that populates them separately.
    """
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        cleaned = raw.strip().strip("`")
        if cleaned.lower().startswith("json"):
            cleaned = cleaned[4:].strip()
        try:
            parsed = json.loads(cleaned)
        except json.JSONDecodeError:
            print("  WARN: risk-factor LLM returned non-JSON — dropping bullets")
            return []
    if not isinstance(parsed, dict):
        print("  WARN: risk-factor LLM response was not a JSON object — dropping bullets")
        return []
    items = parsed.get("executiveSummary")
    if not isinstance(items, list):
        print("  WARN: risk-factor LLM response missing 'executiveSummary' list — dropping bullets")
        return []
    bullets = [b for b in (_normalize_risk_factor_bullet(i) for i in items) if b is not None]
    if not bullets:
        print("  WARN: risk-factor LLM returned 0 valid bullets — dropping bullets")
    return bullets


def _call_risk_factor_azure(
    system_prompt: str,
    user_prompt: str,
    *,
    endpoint: str,
    deployment: str,
    api_key: str,
    api_version: str,
) -> List[Dict[str, str]]:
    """Call Azure OpenAI for risk-factor bullets. Returns [] on any error."""
    try:
        from openai import AzureOpenAI  # type: ignore
    except ImportError:
        print("  WARN: openai package not installed — skipping risk_factor_bullets")
        return []
    try:
        client = AzureOpenAI(
            api_version=api_version,
            azure_endpoint=endpoint,
            api_key=api_key,
            timeout=60.0,
        )
        print(f"  Calling Azure OpenAI for risk factors ({endpoint}, deployment={deployment})…")
        resp = client.chat.completions.create(
            model=deployment,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            temperature=0.3,
            response_format={"type": "json_object"},
            max_tokens=4000,
        )
        raw = resp.choices[0].message.content or ""
    except Exception as e:
        print(f"  WARN: risk-factor Azure call failed ({e!r}) — dropping bullets")
        return []
    return _parse_risk_factor_response(raw)


def _call_risk_factor_litellm(
    system_prompt: str,
    user_prompt: str,
    *,
    base_url: str,
    api_key: str,
    model: str,
) -> List[Dict[str, str]]:
    """Call LiteLLM proxy for risk-factor bullets. Returns [] on any error."""
    try:
        from openai import OpenAI  # type: ignore
    except ImportError:
        print("  WARN: openai package not installed — skipping risk_factor_bullets")
        return []
    try:
        client = OpenAI(base_url=base_url, api_key=api_key, timeout=60.0)
        print(f"  Calling LiteLLM proxy for risk factors ({base_url}, model={model})…")
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            temperature=0.3,
            response_format={"type": "json_object"},
            max_tokens=4000,
        )
        raw = resp.choices[0].message.content or ""
    except Exception as e:
        print(f"  WARN: risk-factor LiteLLM call failed ({e!r}) — dropping bullets")
        return []
    return _parse_risk_factor_response(raw)


def build_risk_factor_bullets(snapshot: Dict[str, Any]) -> List[Dict[str, str]]:
    """Generate the 6 Portfolio Risk Factor bullets used by the Executive Summary.

    Provider selection mirrors ``build_ai_insights``:

    1. **Azure OpenAI (primary)** — used when all three of
       ``AZURE_OPENAI_ENDPOINT``, ``AZURE_OPENAI_DEPLOYMENT``,
       ``AZURE_OPENAI_API_KEY`` are set. ``AZURE_OPENAI_API_VERSION``
       defaults to ``2025-01-01-preview`` when unset.

    2. **LiteLLM proxy (fallback)** — used when the Azure vars aren't
       all set but ``AFN_LITELLM_API_KEY`` is. Local-dev only.

    3. **No config** — logs a WARN and returns ``[]`` so the snapshot
       build never fails just because the LLM is unreachable.

    The prompt lives in ``data/prompts/risk-factor-analysis.system.md`` and
    is shared with PR B's ``/api/ai-analysis`` regenerate endpoint. If the
    file is missing this function logs a WARN and returns ``[]``.
    """
    system_prompt = _load_risk_factor_prompt()
    if system_prompt is None:
        return []

    try:
        facts = _build_risk_factor_facts(snapshot)
    except Exception as e:  # defensive: fact assembly must never crash the build
        print(f"  WARN: risk-factor facts assembly failed ({e!r}) — dropping bullets")
        return []

    user_prompt = (
        "Analyze this FHA portfolio data and generate the executive summary and action items:\n\n"
        + facts
    )

    az_endpoint = os.environ.get("AZURE_OPENAI_ENDPOINT")
    az_deployment = os.environ.get("AZURE_OPENAI_DEPLOYMENT")
    az_api_key = os.environ.get("AZURE_OPENAI_API_KEY")
    az_api_version = os.environ.get("AZURE_OPENAI_API_VERSION", "2025-01-01-preview")

    if az_endpoint and az_deployment and az_api_key:
        return _call_risk_factor_azure(
            system_prompt,
            user_prompt,
            endpoint=az_endpoint,
            deployment=az_deployment,
            api_key=az_api_key,
            api_version=az_api_version,
        )

    litellm_key = os.environ.get("AFN_LITELLM_API_KEY")
    if litellm_key:
        base_url = os.environ.get(
            "AFN_LITELLM_BASE_URL", "http://100.120.169.17:4000/v1"
        )
        model = os.environ.get("AFN_LITELLM_INSIGHT_MODEL", "gpt-4o")
        return _call_risk_factor_litellm(
            system_prompt,
            user_prompt,
            base_url=base_url,
            api_key=litellm_key,
            model=model,
        )

    print(
        "  WARN: no AI provider configured for risk_factor_bullets "
        "(set AZURE_OPENAI_* or AFN_LITELLM_API_KEY) — dropping bullets"
    )
    return []


def build_ai_insights(snapshot: Dict[str, Any]) -> List[Dict[str, str]]:
    """Generate 4 AI insights, preferring Azure OpenAI over LiteLLM.

    Provider selection (in order):

    1. **Azure OpenAI (primary)** — used when all four of
       ``AZURE_OPENAI_ENDPOINT``, ``AZURE_OPENAI_DEPLOYMENT``,
       ``AZURE_OPENAI_API_KEY`` are set. ``AZURE_OPENAI_API_VERSION``
       defaults to ``2025-01-01-preview`` when unset. This is the path
       used by the production Container App (public HTTPS, no tailnet).

    2. **LiteLLM proxy (fallback)** — used when the Azure vars aren't
       all set but ``AFN_LITELLM_API_KEY`` is. Preserves the original
       behaviour for local dev environments that still route through the
       AFN LiteLLM proxy on Matt's tailnet.

    3. **No config** — returns the canned ``_FALLBACK_AI_INSIGHTS`` so
       the snapshot build never breaks.
    """
    az_endpoint = os.environ.get("AZURE_OPENAI_ENDPOINT")
    az_deployment = os.environ.get("AZURE_OPENAI_DEPLOYMENT")
    az_api_key = os.environ.get("AZURE_OPENAI_API_KEY")
    az_api_version = os.environ.get("AZURE_OPENAI_API_VERSION", "2025-01-01-preview")

    if az_endpoint and az_deployment and az_api_key:
        return _build_ai_insights_azure(
            snapshot,
            endpoint=az_endpoint,
            deployment=az_deployment,
            api_key=az_api_key,
            api_version=az_api_version,
        )

    litellm_key = os.environ.get("AFN_LITELLM_API_KEY")
    if litellm_key:
        base_url = os.environ.get(
            "AFN_LITELLM_BASE_URL", "http://100.120.169.17:4000/v1"
        )
        model = os.environ.get("AFN_LITELLM_INSIGHT_MODEL", "gpt-4o")
        return _build_ai_insights_litellm(
            snapshot, base_url=base_url, api_key=litellm_key, model=model
        )

    print(
        "  WARN: no AI provider configured "
        "(set AZURE_OPENAI_* or AFN_LITELLM_API_KEY) — skipping AI insights"
    )
    return list(_FALLBACK_AI_INSIGHTS)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("period", help="YYYY-MM period, e.g. 2026-02")
    ap.add_argument("--out", help="Override output path", default=None)
    args = ap.parse_args()

    period = args.period
    src = SOURCE_ROOT / period
    if not src.is_dir():
        print(f"ERROR: {src} does not exist", file=sys.stderr)
        return 2

    # Pretty "February 2026"
    try:
        year, month = period.split("-")
        label = f"{MONTH_NAMES[int(month) - 1]} {year}"
    except Exception:
        label = period

    print(f"Building snapshot for {period} ({label})")

    # ── Source files ──
    # Slot keys resolve through SLOT_ALIAS_TABLE (see _find_source) so we
    # tolerate every naming convention the RPA — or a human — has ever used,
    # including legacy ``.xls`` uploads (auto-converted to ``.xlsx``).
    total_path = _find_source(period, "hud_total_compare_ratios")
    hoc_path = _find_source(period, "hoc_compare_ratios")
    field_path = _find_source(period, "hud_field_offices")
    branches_path = _find_source(period, "hud_branches")
    nw2_path = _find_source(period, "nw_data")
    enc_candidates = sorted(src.glob("Neighborhood Watch Report*Enc Data*.xlsx")) \
        or sorted(src.glob("*Enc Data*.xlsx"))
    if not enc_candidates:
        raise FileNotFoundError(f"No Encompass export found in {src}")
    enc_path = enc_candidates[0]

    print(f"  Total: {total_path.name}")
    print(f"  HOC:   {hoc_path.name}")
    print(f"  Field: {field_path.name}")
    print(f"  Branch: {branches_path.name}")
    print(f"  NW2:   {nw2_path.name}")
    print(f"  Enc:   {enc_path.name}")

    # ── Compare ratios ──
    print("Reading compare_ratios_total…")
    total_rows, perf_date, perf_label = read_compare_ratios_total(total_path)
    if not perf_date:
        # Fall back to last day of period month
        y, m = period.split("-")
        last_day = (dt.date(int(y), int(m) % 12 + 1, 1) - dt.timedelta(days=1)) if int(m) < 12 \
            else dt.date(int(y), 12, 31)
        perf_date = last_day.isoformat()
    print(f"  Performance period: {perf_date} — {perf_label}")

    print("Reading compare_ratios_hoc…")
    hoc_rows = read_compare_ratios_hoc(hoc_path)
    print(f"  {len(hoc_rows)} HOC rows")

    print("Reading compare_ratios_hud_office…")
    field_rows = read_compare_ratios_hud_office(field_path)
    print(f"  {len(field_rows)} HUD office rows")
    hud_office_lookup = {r["hud_office"]: r for r in field_rows}

    print("Reading compare_ratios_branch…")
    branch_rows = read_compare_ratios_branch(branches_path)
    print(f"  {len(branch_rows)} branch rows")

    # ── HUD total population (optional) ──
    # HUD delivers `NW Total Population <M.D.YY>.xlsx` on request with per-loan
    # detail for AFN's entire endorsed book (not just SDQ). When present, we
    # use it to filter the Encompass loan set to only HUD-endorsed loans so
    # our totals match HUD Compare Ratios exactly. When absent, we fall back
    # to the raw Encompass set (older behavior).
    hud_pop_path: Optional[Path] = None
    try:
        hud_pop_path = _find_source(period, "hud_total_population")
    except FileNotFoundError:
        pass
    hud_pop_by_case: Dict[str, dict] = {}
    if hud_pop_path is not None:
        print(f"  HUD Pop: {hud_pop_path.name}")
        hud_pop_by_case = read_hud_total_population(hud_pop_path)
        print(f"  HUD total population: {len(hud_pop_by_case):,} endorsed loans")

    # ── Loans (Encompass + NW2) ──
    print("Reading loan-level data…")
    loans = build_loans(enc_path, nw2_path, hud_office_lookup)
    print(f"  {len(loans):,} loans from Encompass")

    # ── HUD endorsement filter ──
    # If we have HUD's total-population file, drop any Encompass loan whose
    # FHA case number isn't in HUD's endorsed set. Also annotate each surviving
    # loan with HUD's channel classification (hud_channel) and endorsement flag
    # (hud_endorsed=True) so downstream reporting can compare / drift-check.
    if hud_pop_by_case:
        before = len(loans)
        kept = []
        dropped_no_case = 0
        dropped_not_endorsed = 0
        for l in loans:
            case = l.get("fha_case_number")
            if not case:
                dropped_no_case += 1
                continue
            hud = hud_pop_by_case.get(case)
            if hud is None:
                dropped_not_endorsed += 1
                continue
            l["hud_endorsed"] = True
            l["hud_channel"] = hud["hud_channel"]
            l["hud_orig_id"] = hud["hud_orig_id"]
            l["hud_sponsor_id"] = hud["hud_sponsor_id"]
            l["hud_fha_ins_stat"] = hud["hud_fha_ins_stat"]
            kept.append(l)
        loans = kept
        missing_in_enc = len(hud_pop_by_case) - len(loans)
        print(f"  HUD endorsement filter: {before:,} → {len(loans):,} loans "
              f"(dropped {dropped_not_endorsed:,} not-in-HUD, "
              f"{dropped_no_case:,} no-case-number; HUD has {missing_in_enc:,} "
              f"loans not present in Encompass)")

    # ── Derived aggregates ──
    print("Computing portfolio_slices…")
    slices = build_portfolio_slices(loans)
    print(f"  {len(slices)} slice rows")

    print("Computing loan_officer_performance…")
    lo_perf = build_loan_officer_performance(loans)
    print(f"  {len(lo_perf)} LOs")

    print("Computing risk_indicator_distribution…")
    risk_dist = build_risk_indicator_distribution(loans)

    print("Computing underwriter_rollup…")
    underwriter_rollup = build_underwriter_rollup(loans)
    print(f"  {len(underwriter_rollup)} underwriters")

    print("Computing delinquency_reason_rollup…")
    delinquency_reason_rollup = build_delinquency_reason_rollup(loans)
    print(f"  {len(delinquency_reason_rollup)} reason buckets")

    print("Computing indemnification_loans…")
    indemnification_loans = build_indemnification_loans(loans)
    print(f"  {len(indemnification_loans)} indemnified loans")

    print("Computing sponsor_tpo_detail…")
    sponsor_tpo_detail = build_sponsor_tpo_detail(loans)
    print(f"  {len(sponsor_tpo_detail)} sponsored originators")

    # ── Compose ──
    snapshot = OrderedDict()
    # Top-level schema fields consumed by the container's write_snapshot_outputs()
    # (infra/snapshot-pipeline/container/app.py L461–465) when it builds
    # index.json. Keep in sync with snapshot_meta below — same values, but the
    # container reads the top-level keys, not the nested meta block.
    snapshot["period"] = period
    snapshot["label"] = label
    snapshot["performance_period"] = perf_date
    snapshot["snapshot_meta"] = {
        "period": period,
        "label": label,
        "performance_period": perf_date,
        "performance_period_label": perf_label or label,
        "generated_at": dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
        "generated_by": f"scripts/build-snapshot.py v{SCRIPT_VERSION}",
        "source_files": [
            total_path.name, hoc_path.name, field_path.name,
            branches_path.name, nw2_path.name, enc_path.name,
        ],
        "schema_version": SCHEMA_VERSION,
        "notes": f"{len(loans):,} loans; {len(slices)} portfolio slices; {len(lo_perf)} LOs",
    }
    snapshot["compare_ratios_total"] = total_rows
    snapshot["compare_ratios_hoc"] = hoc_rows
    snapshot["compare_ratios_hud_office"] = field_rows
    # ── Enrich branch rows with AFN names + HUD offices via case-number bridge ──
    print("Enriching HUD branch rows with AFN branch names…")
    _enrich_branch_rows(branch_rows, loans, nw2_path)
    snapshot["compare_ratios_branch"] = branch_rows
    snapshot["portfolio_slices"] = slices
    snapshot["loan_officer_performance"] = lo_perf
    snapshot["risk_indicator_distribution"] = risk_dist
    snapshot["underwriter_rollup"] = underwriter_rollup
    snapshot["delinquency_reason_rollup"] = delinquency_reason_rollup
    snapshot["indemnification_loans"] = indemnification_loans
    snapshot["sponsor_tpo_detail"] = sponsor_tpo_detail
    snapshot["loans"] = loans

    # ── Projections ─ loan-level → office/HOC/national aggregation ──
    # Must run before ai_insights so the LLM prompt can see projected data.
    print("Computing projections (1/3/6mo × best/base/worst)…")
    try:
        from build_projections import build_projections  # sibling module
    except ImportError:
        # Fallback: script may be invoked from a working directory outside
        # scripts/. Explicitly add the scripts dir to sys.path and retry.
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        from build_projections import build_projections  # type: ignore
    snapshot["projections"] = build_projections(snapshot)
    _proj_n_offices = len(snapshot["projections"]["offices"])
    _proj_n_crossings = sum(
        len(o["threshold_crossings"]) for o in snapshot["projections"]["offices"]
    )
    print(
        f"  {_proj_n_offices} offices projected, "
        f"{_proj_n_crossings} threshold-crossing events"
    )

    # ── AI insights (LLM-generated narrative findings) ──
    print("Generating AI insights…")
    snapshot["ai_insights"] = build_ai_insights(snapshot)
    print(f"  {len(snapshot['ai_insights'])} insight(s) produced")

    # ── Portfolio Risk Factor bullets (baked at build time; see PR A/B) ──
    # Frontend Executive Summary card historically fetched these via
    # /api/ai-analysis on every period load. Baking them into the snapshot
    # removes that per-load cost. PR B introduces the on-demand "regenerate"
    # write-back path that mutates the ``regenerated_*`` fields below.
    print("Generating Portfolio Risk Factor bullets…")
    _rfb_generated_by = snapshot["snapshot_meta"]["generated_by"]
    _rfb_bullets = build_risk_factor_bullets(snapshot)
    snapshot["risk_factor_bullets"] = {
        "bullets": _rfb_bullets,
        "generated_at": snapshot["snapshot_meta"]["generated_at"],
        "generated_by": _rfb_generated_by,
        "regenerated_by": None,
        "regenerated_at": None,
        "schema_version": RISK_FACTOR_BULLETS_SCHEMA_VERSION,
    }
    print(f"  {len(_rfb_bullets)} risk-factor bullet(s) baked")

    # ── Write ──
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = Path(args.out) if args.out else SNAPSHOT_DIR / f"{period}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(snapshot, f, indent=2, default=str)
    size_mb = out_path.stat().st_size / (1024 * 1024)
    print(f"Wrote {out_path} ({size_mb:.2f} MB)")

    # ── Update index.json ──
    index_path = SNAPSHOT_DIR / "index.json"
    if index_path.exists():
        try:
            index = json.loads(index_path.read_text(encoding="utf-8"))
        except Exception:
            index = {"periods": [], "updated_at": "", "schema_version": SCHEMA_VERSION}
    else:
        index = {"periods": [], "updated_at": "", "schema_version": SCHEMA_VERSION}

    periods = [p for p in index.get("periods", []) if p.get("period") != period]
    periods.append({
        "period": period,
        "label": label,
        "performance_period": perf_date,
        "generated_at": snapshot["snapshot_meta"]["generated_at"],
        "file": out_path.name,
    })
    periods.sort(key=lambda p: p["period"], reverse=True)
    index["periods"] = periods
    index["updated_at"] = dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    index["schema_version"] = SCHEMA_VERSION
    with open(index_path, "w", encoding="utf-8") as f:
        json.dump(index, f, indent=2)
    print(f"Updated {index_path} — {len(periods)} period(s) indexed")

    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
